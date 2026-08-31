# ============================================
# KARBONAT P2 - GSTC / TGA Uyumlu Karbon Ayak İzi
# Akıllı Sürdürülebilirlik Raporlaması
# v0.4 - Yeniden Tasarım
# ============================================

import os
import pandas as pd
from datetime import datetime
from io import BytesIO

import streamlit as st

from factors import EMISSION_FACTORS, KATEGORI_ACIKLAMALARI
from emission_calc import (
    hesapla_scope_ayrimi,
    hesapla_normalize_metrikler,
    en_agir_kaynaklar,
)
from tga_tables import (
    format_donem,
    tum_tablolar,
)
from green_report import save_green_report
from icerik_hub import (
    AMAC_GRUPLARI,
    ICERIK_TURLERI,
    ISKELETLER,
    TEMALAR,
    tercih_sorulari,
    varsayilan_tercih,
)
import ai_engine
import tasarim
from data_store import (
    list_facilities,
    get_facility,
    save_facility,
    list_records,
    get_record,
    get_previous_record,
    save_record,
    create_user,
    verify_login,
    get_user_by_id,
    count_facilities,
    get_content_prefs,
    save_content_prefs,
    get_report,
    save_report,
    get_media,
    save_media,
    _db,
    _save,
)
import raporlar

# ==============================================
# MARKA KİMLİĞİ - KARBON & ORMAN
# ==============================================
INK      = "#17211b"   # grafit (karbon)
PRIMARY  = "#1d6b45"   # zümrüt
PRIMARY2 = "#2e8b57"   # orman
ACCENT   = "#3da873"   # yaprak
SAGE     = "#b7d8c2"   # açık adaçayı
MIST     = "#eef4ee"   # sisli yeşil
CARD     = "#ffffff"
PAPER    = "#f6f8f4"
BORDER   = "#e3e9e3"
MUTED    = "#6b7a70"
AMBER    = "#d99a3d"   # öne çıkan
TERRA    = "#c0573d"

st.set_page_config(
    page_title="KarbonAT P2 - Otel Karbon Ayak İzi",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==============================================
# PREMIUM STİL SİSTEMİ
# ==============================================
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@500;600;700;800&family=JetBrains+Mono:wght@500&display=swap');
    :root {{
        --ink: {INK}; --primary: {PRIMARY}; --primary2:{PRIMARY2}; --accent:{ACCENT};
        --sage:{SAGE}; --mist:{MIST}; --card:{CARD}; --paper:{PAPER}; --border:{BORDER}; --muted:{MUTED};
        --amber:{AMBER}; --terra:{TERRA};
        --r-sm: 12px; --r-md: 16px; --r-lg: 22px; --r-xl: 28px;
        --shadow-sm: 0 2px 10px rgba(23,33,27,.06);
        --shadow-md: 0 8px 28px rgba(23,33,27,.08);
        --shadow-lg: 0 18px 44px rgba(23,33,27,.13);
    }}
    .stApp {{ background:
        radial-gradient(900px 500px at 12% -6%, rgba(61,168,115,.12), transparent 60%),
        radial-gradient(800px 600px at 92% 0%, rgba(217,154,61,.10), transparent 55%),
        linear-gradient(180deg, #f8faf7 0%, var(--paper) 100%);
    }}
    html, body, p, li, span, div, label {{ color: var(--ink); font-family: 'Plus Jakarta Sans','Segoe UI',system-ui,sans-serif; }}
    h1,h2,h3 {{ font-family: 'Plus Jakarta Sans', sans-serif; letter-spacing: -0.6px; }}
    h1 {{ font-size: 32px; font-weight: 800; }}
    [data-testid="stSidebar"] {{
        background: rgba(255,255,255,.86); backdrop-filter: blur(14px);
        border-right: 1px solid rgba(227,233,227,.95); box-shadow: 6px 0 24px rgba(23,33,27,.04);
    }}
    [data-testid="stSidebarNav"] {{ display:none; }}
    header[data-testid="stHeader"] {{ background: rgba(246,248,244,.82); backdrop-filter: blur(10px); border-bottom: 1px solid rgba(227,233,227,.65); }}
    header[data-testid="stHeader"] > div {{ background: transparent !important; }}
    .block-container {{ padding-top: 3.2rem; max-width: 1220px; padding-bottom: 2.2rem; }}
    /* hero asla header altinda kesilmesin */
    .hero {{ margin-top: 6px; scroll-margin-top: 72px; }}
    /* scroll bar */
    ::-webkit-scrollbar {{ width:10px; height:10px; }}
    ::-webkit-scrollbar-thumb {{ background: #d7e2d7; border-radius: 999px; border: 2px solid #eef4ee; }}
    /* SIDEBAR */
    .sb-brand {{ font-size:22px; font-weight:800; letter-spacing:-0.5px; color:var(--primary); }}
    .sb-brand small {{ font-weight:700; color:var(--amber); margin-left:2px; }}
    .sb-tag {{ font-size:11px; color:var(--muted); letter-spacing:.14em; text-transform:uppercase; margin: 2px 0 14px; }}
    .sb-card {{
        background: linear-gradient(180deg, #ffffff 0%, #f7faf7 100%);
        border:1px solid var(--border); border-radius: var(--r-md);
        padding:14px 16px; margin:10px 0; box-shadow: var(--shadow-sm);
        position:relative; overflow:hidden;
    }}
    .sb-card::before {{ content:""; position:absolute; left:0; top:0; bottom:0; width:3px; background: linear-gradient(180deg, var(--primary), var(--accent)); opacity:.9; }}
    .sb-label {{ font-size:11px; text-transform:uppercase; letter-spacing:.1em; color:var(--muted); font-weight:700; margin-bottom:8px; }}
    .sb-value {{ font-weight:800; font-size:15px; }}
    /* step nav as pills */
    div[data-testid="stSidebar"] .stButton > button {{
        text-align:left; justify-content:flex-start; border-radius: 12px; padding:10px 14px; font-size:14px;
        background: transparent !important; color: var(--muted) !important; border:1px solid transparent !important; box-shadow:none !important;
    }}
    div[data-testid="stSidebar"] .stButton > button:hover {{ background: var(--mist) !important; color: var(--primary) !important; }}
    div[data-testid="stSidebar"] .stButton > button:disabled {{ opacity:.45; }}
    /* global buttons */
    .stButton > button {{
        background: linear-gradient(180deg, var(--primary) 0%, #17613f 100%); color:#fff; border:none;
        padding: 12px 22px; border-radius: 14px; font-weight:800; font-size:14.5px; letter-spacing:-0.2px;
        box-shadow: 0 8px 18px rgba(23,107,69,.22); transition: transform .16s, box-shadow .16s, filter .16s;
    }}
    .stButton > button:hover {{ transform: translateY(-1px); box-shadow: 0 10px 22px rgba(23,107,69,.26); filter: brightness(1.02); }}
    .stButton > button:active {{ transform: translateY(0px); }}
    button[kind="secondary"] {{
        background:#fff !important; color: var(--primary) !important; border:1.5px solid #cfe3d4 !important;
        box-shadow: var(--shadow-sm) !important;
    }}
    button[kind="secondary"]:hover {{ border-color: var(--primary) !important; background:#f6faf7 !important; }}
    .stDownloadButton > button {{
        background: linear-gradient(180deg, #1e6e46, #144a2e); color:#fff; border-radius:14px; font-weight:800;
        box-shadow: 0 6px 14px rgba(23,107,69,.18);
    }}
    /* KPI */
    .kpi-grid {{ display:grid; grid-template-columns: repeat(4,1fr); gap:14px; margin:14px 0; }}
    @media (max-width: 1100px) {{ .kpi-grid {{ grid-template-columns: repeat(2,1fr); }} }}
    @media (max-width: 640px) {{ .kpi-grid {{ grid-template-columns: 1fr; }} }}
    .kpi-card {{
        background: linear-gradient(180deg, #ffffff 0%, #fbfdfb 100%); border:1px solid var(--border);
        border-radius: var(--r-lg); padding:18px 18px; box-shadow: var(--shadow-sm); position:relative; overflow:hidden;
    }}
    .kpi-card::after {{ content:""; position:absolute; top:0; left:0; right:0; height:3px; background: linear-gradient(90deg, var(--primary), var(--accent)); opacity:.9; }}
    .kpi-label {{ font-size:11px; text-transform:uppercase; letter-spacing:.09em; color:var(--muted); font-weight:700; }}
    .kpi-value {{ font-size:24px; font-weight:800; letter-spacing:-0.6px; }}
    .kpi-sub {{ font-size:12.5px; color:var(--muted); }}
    .data-card {{
        background: rgba(255,255,255,.92); border:1px solid var(--border); border-radius: var(--r-lg);
        padding:18px 18px; box-shadow: var(--shadow-sm); backdrop-filter: blur(6px);
    }}
    .banner {{
        background: linear-gradient(180deg, #ffffff, #f0f7f1); border:1px solid #dbe8de; border-left:4px solid var(--accent);
        border-radius:14px; padding:14px 16px; color: var(--ink); box-shadow: var(--shadow-sm);
    }}
    .chip {{
        background: #fff; color: var(--primary); border:1px solid #d8e8dc;
        padding:6px 12px; border-radius:999px; font-weight:800; font-size:12.5px; box-shadow: var(--shadow-sm);
    }}
    .hero {{
        text-align:center; padding: 38px 18px 22px; position:relative; overflow:hidden;
        background: radial-gradient(700px 320px at 50% -10%, rgba(61,168,115,.16), transparent 70%),
                    linear-gradient(180deg, rgba(255,255,255,.96), rgba(238,244,238,.9));
        border:1px solid rgba(227,233,227,.9); border-radius: 22px; box-shadow: var(--shadow-md);
        margin: 6px 0 18px;
    }}
    .hero .logo {{ font-size:52px; filter: drop-shadow(0 6px 16px rgba(23,107,69,.18)); }}
    .hero h1 {{ font-size:38px; margin:6px 0 4px; }}
    .hero p {{ color: var(--muted); max-width:640px; margin:0 auto; font-size:15.5px; line-height:1.5; }}
    .section-title {{ font-size:16px; font-weight:800; letter-spacing:-0.3px; margin:22px 0 10px; display:flex; align-items:center; gap:10px; }}
    .section-title::before {{ content:""; width:28px; height:3px; border-radius:999px; background: linear-gradient(90deg, var(--primary), var(--accent)); display:inline-block; }}
    .footer {{ text-align:center; color:var(--muted); font-size:12.5px; padding:34px 0 10px; }}
    /* form */
    .stTextInput input, .stNumberInput input {{ border-radius:12px; border:1px solid #dbe3db; background:#fff; }}
    .stTextInput input:focus, .stNumberInput input:focus {{ border-color: var(--accent); box-shadow: 0 0 0 3px rgba(61,168,115,.14); }}
    .stExpander {{ background: rgba(255,255,255,.94); border:1px solid var(--border); border-radius:16px; box-shadow: var(--shadow-sm); }}
    .stTabs [data-baseweb="tab-list"] {{ gap:8px; background: rgba(255,255,255,.72); padding:6px; border-radius: 14px; border:1px solid var(--border); }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 6px !important; background: #f1f5f9 !important; padding: 6px !important; border-radius: 12px !important; border: 1px solid #e2e8f0 !important; display:inline-flex !important; flex-wrap:wrap !important; margin-bottom: 16px !important; }}
    .stTabs [data-baseweb="tab"] {{
        background: transparent !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 9px 16px !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        color: #64748b !important;
        transition: all 0.18s ease !important;
    }}
    .stTabs [data-baseweb="tab"]:hover {{
        background: #ffffff !important;
        color: #0f172a !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
    }}
    .stTabs [aria-selected="true"] {{
        background: #166534 !important;
        color: #ffffff !important;
        box-shadow: 0 2px 8px rgba(22, 101, 52, 0.20) !important;
    }}
    .tga-card {{ background:#ffffff; border:1px solid #e5e7eb; border-radius:12px; padding:16px; margin:8px 0; box-shadow:0 1px 3px rgba(0,0,0,0.04); transition: all .18s ease; }}
    .tga-card:hover {{ border-color:#bbf7d0; box-shadow:0 4px 12px rgba(22,101,52,0.08); transform: translateY(-1px); }}
    .tga-card:focus-within {{ border-color:#166534; box-shadow:0 0 0 3px rgba(22,101,52,0.10); }}
    div[data-testid="stNumberInput"] input {{
        border: 1px solid #d1d5db !important;
        border-radius: 6px !important;
        background: #ffffff !important;
        font-size: 14px !important;
        color: #0f172a !important;
    }}
    div[data-testid="stNumberInput"] input:focus {{
        border-color: #166534 !important;
        box-shadow: 0 0 0 3px rgba(22, 101, 52, 0.12) !important;
    }}
    .stTabs [aria-selected="true"] {{ background: var(--primary); color:#fff !important; box-shadow: var(--shadow-sm); }}
    [data-testid="stMetricValue"] {{ font-weight:800; color:var(--primary); }}
    /* data card hover */
    .data-card:hover {{ border-color:#d4e3d6; box-shadow: var(--shadow-md); transform: translateY(-1px); transition:.18s; }}
    /* MVP Stepper - horizontal */
    .mvp-stepper {{display:flex; align-items:center; gap:8px; flex-wrap:wrap; margin:10px 0 14px; padding:8px; background:#fff; border:1px solid var(--border); border-radius:12px; box-shadow: var(--shadow-sm);}}
    .mvp-step {{display:flex; align-items:center; gap:8px; padding:6px 12px; border-radius:999px; font-size:13px; font-weight:700; color:var(--muted); background:transparent; border:1px solid transparent;}}
    .mvp-step.active {{background:var(--primary); color:#fff; border-color:var(--primary); box-shadow:0 2px 8px rgba(29,107,69,.18);}}
    .mvp-step.done {{background:var(--mist); color:var(--primary); border-color:var(--sage);}}
    .mvp-step .num {{width:22px; height:22px; border-radius:50%; display:inline-flex; align-items:center; justify-content:center; font-size:12px; font-weight:800; background:var(--border); color:var(--muted);}}
    .mvp-step.active .num {{background:#fff; color:var(--primary);}}
    .mvp-step.done .num {{background:var(--sage); color:var(--primary);}}
    .mvp-arrow {{color:var(--muted); font-weight:800;}}
    .mvp-cta {{position:sticky; bottom:12px; z-index:20; background:rgba(255,255,255,.96); backdrop-filter:blur(8px); border:1px solid var(--border); border-radius:14px; padding:12px; box-shadow: var(--shadow-md); margin-top:18px;}}
    .mvp-empty {{background: linear-gradient(180deg, #fff, #f7faf7); border:1px solid var(--border); border-radius:14px; padding:22px; text-align:center; box-shadow: var(--shadow-sm);}}
</style>
""", unsafe_allow_html=True)


# ==============================================
# OTURUM
# ==============================================
def init_session():
    defaults = {
        "user": None,
        "step": 0,
        "facility_id": None,
        "tesis": {},
        "period": datetime.today().strftime("%Y-%m"),
        "musteri": 0,
        "dolu_oda_gun": 0,
        "tuketim": {},
        "sonuc": None,
        "history": [],
        "kategori_totallari": {},
        "atik_bertaraf": "Geri dönüşüm + çöp (karışık)",
        "yenilenebilir_oran": 30,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def reset_to_home():
    """Tüm tesis/veri state'ini temizle ve anasayfaya (step 0) dön."""
    # Tüm facility/veri state'ini topla sil
    keys_to_clear = [
        "step", "facility_id", "tesis", "tuketim", "sonuc", "history",
        "kategori_totallari", "musteri", "dolu_oda_gun",
        "atik_bertaraf", "yenilenebilir_oran", "_load_period",
        "show_raporlar", "categori_totallari",
    ]
    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]
    # Widget key'leri de temizle (veri_*, rte_, rmod_, vb.)
    widget_keys = [k for k in st.session_state.keys()
                   if k.startswith(("veri_", "rte_", "rmod_", "rkapat_",
                                    "rug_", "rapdf_", "radoc_", "raex_",
                                    "rasv_", "rkcd_", "rapor_detay_",
                                    "medya_detay_", "mkapat_", "mte_",
                                    "tp_", "uretim_", "gorsel_",
                                    "mga_", "mpdf_", "mpng_",
                                    "mthtml_", "mdoc_", "msav_",
                                    "mcd_"))]
    for k in widget_keys:
        del st.session_state[k]
    init_session()
    st.session_state.step = 0
    st.rerun()


init_session()

KATEGORI_BIRIM = {
    "Elektrik": "kWh",
    "Doğal Gaz ve Yakıtlar": "m³ / kg / L",
    "Soğutucu & F-Gaz (Scope 1)": "kg",
    "Araç Filosu & İş Seyahatleri (Scope 1/3)": "L / km",
    "Su": "m³",
    "Gıda Tüketimi": "kg",
    "Atık Yönetimi": "kg",
    "Kimyasal Tüketimi": "L",
}

STEP_ISIMLERI = [
    ("🏠", "Tesis"),
    ("🏨", "Profil"),
    ("📊", "Veri Girişi"),
    ("🧮", "Hesaplama"),
    ("📄", "Rapor"),
    ("📣", "Medya"),
]


# ==============================================
# SIDEBAR
# ==============================================
def _clear_facility_state():
    """Tesis değişince veri/sonuç state'ini temizle (widget key'leri de)."""
    keys_to_clear = [
        "tesis", "tuketim", "sonuc", "history",
        "kategori_totallari", "musteri", "dolu_oda_gun",
        "atik_bertaraf", "yenilenebilir_oran", "_load_period",
        "show_raporlar",
    ]
    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]
    # Widget key'leri (veri_*, rte_*, vb.) de temizle
    widget_keys = [k for k in st.session_state.keys()
                   if k.startswith(("veri_", "rte_", "rmod_", "rkapat_",
                                    "rug_", "rapdf_", "radoc_", "raex_",
                                    "rasv_", "rkcd_", "rapor_detay_",
                                    "medya_detay_", "mkapat_", "mte_",
                                    "tp_", "uretim_", "gorsel_",
                                    "mga_", "mpdf_", "mpng_",
                                    "mthtml_", "mdoc_", "msav_",
                                    "mcd_"))]
    for k in widget_keys:
        del st.session_state[k]


def _save_draft(fac_id):
    """Mevcut form verisini taslak olarak kaydet (hesaplamadan)."""
    draft = {
        "fac_id": fac_id,
        "period": st.session_state.period,
        "musteri": st.session_state.musteri,
        "dolu_oda_gun": st.session_state.dolu_oda_gun,
        "tuketim": st.session_state.tuketim,
        "atik_bertaraf": st.session_state.atik_bertaraf,
        "atik_bertaraf_idx": ATIK_BERTARAF_SECENEKLERI.index(st.session_state.atik_bertaraf),
        "yenilenebilir_oran": st.session_state.yenilenebilir_oran,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "is_draft": True,
    }
    # drafts ayrı bir key altında sakla
    db = _db()
    drafts = db.setdefault("drafts", {}).setdefault(fac_id, {})
    drafts[st.session_state.period] = draft
    _save(db)


def _load_previous_period():
    """Önceki dönem kaydını yükle (manuel)."""
    fac_id = st.session_state.facility_id
    current = st.session_state.period
    prev = get_previous_record(fac_id, current)
    if prev:
        st.session_state.tuketim = prev["tuketim"]
        st.session_state.musteri = prev["musteri"]
        st.session_state.dolu_oda_gun = prev["dolu_oda_gun"]
        st.session_state.atik_bertaraf = prev["atik_bertaraf"]
        st.session_state.yenilenebilir_oran = prev.get("yenilenebilir_oran", 30)
        st.session_state._load_period = prev["period"]
        st.toast(f"Önceki dönem ({format_donem(prev['period'])}) yüklendi", icon="🔙")
    else:
        st.warning("Önceki dönem kaydı bulunamadı")


def sidebar():
    with st.sidebar:
        st.markdown('<div class="sb-brand">🌿 KarbonAT <span style="color:#d99a3d;">P2</span></div>'
                    '<div class="sb-tag">GSTC / TGA Uyumlu Raporlama</div>', unsafe_allow_html=True)

        # Tesis bilgisi
        t = st.session_state.tesis
        if t and t.get("ad"):
            st.markdown(
                f'<div class="sb-card">'
                f'<div class="sb-label">Tesis</div>'
                f'<div class="sb-value">{t["ad"]}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div class="sb-card"><div class="sb-label">Tesis</div>'
                f'<div class="sb-value" style="color:#b6beb8;">Seçilmedi</div></div>',
                unsafe_allow_html=True,
            )

        if st.session_state.step >= 2 and st.session_state.tesis:
            st.markdown(
                f'<div class="sb-card"><div class="sb-label">Dönem</div>'
                f'<div class="sb-value">📅 {format_donem(st.session_state.period)}</div></div>',
                unsafe_allow_html=True,
            )

        # Adım göstergesi (tıklanabilir navigasyon)
        st.markdown(f'<div class="sb-label" style="margin-top:14px;">İlerleme</div>', unsafe_allow_html=True)
        step = st.session_state.step
        for i, (emoji, isim) in enumerate(STEP_ISIMLERI):
            cls = "active" if i == step else ("done" if i < step else "")
            disabled = False
            # Sadece tamamlanan/aktif adımlara gitmeye izin ver; ileri atlama yok
            if i > step and step < 5:
                disabled = True
            if st.button(
                f"{emoji} {isim}",
                key=f"sb_nav_{i}",
                width='stretch',
                disabled=disabled,
                help=("Aktif adım" if i == step else
                      ("Tamamlandı — geri dön" if i < step else "Önce önceki adımları tamamlayın")),
            ):
                # State temizliği gerekirse (örn. step 4→2 veri değişirse)
                if i < step:
                    # Geri dönüşte veri state'i koru (sadece step değişir)
                    st.session_state.step = i
                else:
                    st.session_state.step = i
                st.rerun()

        st.markdown("---")
        if st.button("🔄 Sıfırla (Anasayfa)", width='stretch', type="secondary"):
            reset_to_home()

        st.markdown('<div class="sb-label" style="margin-top:20px;">Oturum</div>', unsafe_allow_html=True)
        if st.session_state.user:
            u = st.session_state.user
            st.markdown(
                f'<div class="sb-card">'
                f'<div class="sb-label">👤 Kullanıcı</div>'
                f'<div class="sb-value">{u.get("fullname", u["username"])}</div>'
                f'<div style="font-size:12px; color:#6b7a70;">@{u["username"]} · {count_facilities(u["id"])} otel</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
            if st.button("🚪 Çıkış Yap", width='stretch', type="secondary"):
                # Kullanıcı + tüm facility state'i temizle
                all_keys = list(st.session_state.keys())
                for k in all_keys:
                    del st.session_state[k]
                init_session()
                st.rerun()


def mvp_stepper(active=0):
    steps = [("1","Tesis"),("2","Veri"),("3","Hesap"),("4","Rapor"),("5","Medya")]
    html = '<div class="mvp-stepper">'
    for i,(n,lbl) in enumerate(steps):
        cls = "active" if i==active else ("done" if i<active else "")
        html += f'<div class="mvp-step {cls}"><span class="num">{n}</span> {lbl}</div>'
        if i < len(steps)-1:
            html += '<span class="mvp-arrow">→</span>'
    html += '</div>'
    return html


def _ai_fallback_icerik(tur_id, tesis, sonuc):
    """AI kotası doluyken deterministik taslak üretir (boş ekran yerine)."""
    from icerik_hub import ISKELETLER
    iskelet = ISKELETLER.get(tur_id, [])
    ad = tesis.get("ad", "Tesisimiz")
    m = (sonuc or {}).get("metrikler", {}) if isinstance(sonuc, dict) else {}
    ton = m.get("toplam_ton", "?")
    lines = [f"> _Not: AI kotası dolu — deterministik taslak gösteriliyor. Daha zengin metin için birkaç dakika sonra tekrar deneyin._", ""]
    for baslik in iskelet:
        lines.append(f"### {baslik}")
        if "ton" in ton.__str__().lower() or baslik.lower().startswith("kapak"):
            lines.append(f"{ad} için {baslik.lower()} — hesaplanan toplam {ton} ton CO₂e verisine dayalı taslak metin. Detaylar için veri girişini düzenleyin.")
        else:
            lines.append(f"{ad} — {baslik} için taslak içerik (AI kapalıyken gösterilen yedek metin).")
        lines.append("")
    return "\n".join(lines)


def _ai_fallback_rapor(sab, tesis, sonuc):
    """Rapor için deterministik fallback."""
    ad = tesis.get("ad", "Tesisimiz")
    m = (sonuc or {}).get("metrikler", {}) if isinstance(sonuc, dict) else {}
    ton = m.get("toplam_ton", "?")
    cols = ", ".join(sab.get("cikti", []))
    return f"> _AI kotası dolu — deterministik taslak. {sab['baslik']} şablonu ({cols}) için yer tutucu._\n\n### {sab['baslik']}\n\n{ad} — {ton} ton CO₂e hesap verisine dayalı rapor taslağı. Gerçek AI üretimi için birkaç dakika sonra Yenile'ye basın.\n\n| Alan | Değer |\n|---|---|\n| Tesis | {ad} |\n| Toplam | {ton} ton CO₂e |\n"

# ==============================================
# YARDIMCILAR
# ==============================================
def _period_options():
    options = []
    now = datetime.today()
    y, m = now.year, now.month
    for _ in range(24):
        options.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return options


def _sifir_tuketim():
    tuk = {}
    for kategori in EMISSION_FACTORS:
        tuk[kategori] = {}
        for alt_tur in EMISSION_FACTORS[kategori]:
            tuk[kategori][alt_tur] = 0.0
    return tuk


def _kategori_df(kategori, tuketim):
    rows = []
    for alt_tur, faktor in EMISSION_FACTORS.get(kategori, {}).items():
        miktar = tuketim.get(kategori, {}).get(alt_tur, 0.0)
        rows.append({"Alt Tür": alt_tur, "Miktar": round(miktar, 2),
                     "Emisyon Faktörü": faktor, "Emisyon (kg)": round(miktar * faktor, 2)})
    return pd.DataFrame(rows)


def _tesis_ozet(t):
    return f'{t["m2"]} m² · {t["oda"]} oda · {t["personel"]} personel'


def _hesapla(tesis, tuketim, musteri, dolu_oda_gun, atik_bertaraf="", yenilenebilir_oran=None):
    y_oran = yenilenebilir_oran if yenilenebilir_oran is not None else 0
    scope_data = hesapla_scope_ayrimi(tuketim, y_oran)
    metrikler = hesapla_normalize_metrikler(
        scope_data["toplam"], tesis["m2"], tesis["oda"], tesis["personel"],
        musteri, dolu_oda_gun,
    )
    metrikler["scope1_kg"] = round(scope_data["scope1"], 2)
    metrikler["scope2_kg"] = round(scope_data["scope2"], 2)
    metrikler["scope3_kg"] = round(scope_data["scope3"], 2)
    en_agir = en_agir_kaynaklar(scope_data["kategori_toplamlari"])
    return {
        "tesis": {**tesis, "musteri": musteri, "dolu_oda_gun": dolu_oda_gun},
        "statik": {
            "atik_bertaraf": atik_bertaraf or tesis.get("atik_bertaraf", ""),
            "yenilenebilir": (
                tesis.get("yenilenebilir")
                if yenilenebilir_oran is None else yenilenebilir_oran
            ),
        },
        "tuketim": tuketim,
        "scope": scope_data,
        "metrikler": metrikler,
        "en_agir": en_agir,
    }


def _color_for_pct(pct):
    if pct > 40:
        return TERRA
    elif pct > 20:
        return PRIMARY
    elif pct > 10:
        return PRIMARY2
    else:
        return SAGE


# ==============================================
# GİRİŞ / KAYIT
# ==============================================
def auth_screen():
    st.markdown('''
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        /* Hide all Streamlit Chrome & Headers */
        #MainMenu, header, footer, [data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stDecoration"], [data-testid="stStatusWidget"] {
            display: none !important;
        }
        
        /* Force full screen viewport without top padding or margins */
        html, body, .stApp, .main, [data-testid="stAppViewContainer"], [data-testid="stAppViewBlockContainer"] {
            padding: 0 !important;
            margin: 0 !important;
            max-width: 100% !important;
            width: 100% !important;
            background-color: #ffffff !important;
            font-family: 'Inter', -apple-system, sans-serif !important;
        }

        .block-container {
            padding: 0 !important;
            margin: 0 !important;
            max-width: 100% !important;
            width: 100% !important;
        }

        /* Target Streamlit Horizontal Block (Column Container) */
        div[data-testid="stHorizontalBlock"] {
            min-height: 100vh !important;
            width: 100% !important;
            margin: 0 !important;
            gap: 0 !important;
            display: flex !important;
        }

        /* AGGRESSIVE LEFT COLUMN TARGETING (Dark Green Side) */
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:first-child,
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(1),
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-of-type(1),
        div[data-testid="stColumn"]:first-child {
            background-color: #0f3822 !important;
            background: #0f3822 !important;
            color: #ffffff !important;
            padding: 60px 48px !important;
            min-height: 100vh !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
            flex: 1 1 50% !important;
            box-sizing: border-box !important;
        }

        /* Force EVERY element in left column to be white */
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:first-child *,
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(1) *,
        div[data-testid="stColumn"]:first-child * {
            color: #ffffff !important;
        }

        /* AGGRESSIVE RIGHT COLUMN TARGETING (Pure White Side) */
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:last-child,
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(2),
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-of-type(2),
        div[data-testid="stColumn"]:last-child {
            background-color: #ffffff !important;
            background: #ffffff !important;
            color: #0f172a !important;
            padding: 60px 48px !important;
            min-height: 100vh !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
            align-items: center !important;
            flex: 1 1 50% !important;
            box-sizing: border-box !important;
        }

        /* Ensure vertical block inside right column is centered and has proper width */
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"] {
            width: 100% !important;
            max-width: 420px !important;
            margin: 0 auto !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
        }

        /* Custom Left Branding Styling */
        .brand-container {
            max-width: 520px;
            margin: 0 auto;
        }
        .brand-kicker {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            border: 1px solid rgba(255,255,255,0.22);
            background: rgba(255,255,255,0.08);
            color: #d8efe1 !important;
            padding: 6px 14px;
            border-radius: 999px;
            font-size: 11px;
            font-weight: 700;
            letter-spacing: 0.08em;
            margin-bottom: 20px;
        }
        .brand-title {
            font-size: 46px;
            font-weight: 800;
            color: #ffffff !important;
            letter-spacing: -1.4px;
            line-height: 1.05;
            margin: 0 0 10px 0;
        }
        .brand-title span {
            color: #86efac !important;
        }
        .brand-sub {
            font-size: 20px;
            font-weight: 600;
            color: #ffffff !important;
            margin: 0 0 12px 0;
        }
        .brand-desc {
            font-size: 14.5px;
            color: #cfe8d6 !important;
            line-height: 1.6;
            margin-bottom: 32px;
        }
        .bullet-item {
            display: flex;
            gap: 16px;
            align-items: flex-start;
            margin: 18px 0;
        }
        .bullet-icon {
            width: 42px;
            height: 42px;
            border-radius: 12px;
            background: rgba(255,255,255,0.12);
            border: 1px solid rgba(255,255,255,0.18);
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 18px;
            flex-shrink: 0;
            color: #ffffff !important;
        }
        .bullet-text b {
            color: #ffffff !important;
            font-size: 15px;
            font-weight: 700;
            display: block;
            margin-bottom: 3px;
        }
        .bullet-text span {
            color: #b7d8c2 !important;
            font-size: 13px;
            line-height: 1.45;
            display: block;
        }

        /* Custom Right Form Styling */
        .auth-head-title {
            font-size: 26px;
            font-weight: 800;
            color: #0f172a !important;
            text-align: center;
            margin: 0 0 6px 0;
            letter-spacing: -0.6px;
        }
        .auth-head-sub {
            font-size: 14px;
            color: #64748b !important;
            text-align: center;
            margin: 0 0 24px 0;
        }

        /* Modern Input Styling */
        div[data-testid="stTextInput"] label {
            font-size: 13px !important;
            font-weight: 600 !important;
            color: #0f172a !important;
            margin-bottom: 6px !important;
        }
        div[data-testid="stTextInput"] input {
            background-color: #f8fafc !important;
            border: 1px solid #e2e8f0 !important;
            border-radius: 8px !important;
            padding: 12px 14px !important;
            font-size: 14px !important;
            color: #0f172a !important;
            box-shadow: none !important;
            transition: all 0.18s ease !important;
        }
        div[data-testid="stTextInput"] input:focus {
            background-color: #ffffff !important;
            border-color: #166534 !important;
            box-shadow: 0 0 0 3px rgba(22, 101, 52, 0.12) !important;
            outline: none !important;
        }

        /* Segmented Control as Sleek Tabs */
        div[data-testid="stSegmentedControl"] {
            background-color: #f1f5f9 !important;
            border-radius: 10px !important;
            padding: 4px !important;
            gap: 4px !important;
            width: 100% !important;
            margin-bottom: 16px !important;
        }
        div[data-testid="stSegmentedControl"] button {
            border-radius: 8px !important;
            font-weight: 600 !important;
            font-size: 13px !important;
            border: none !important;
            background: transparent !important;
            color: #64748b !important;
            flex: 1 !important;
        }
        div[data-testid="stSegmentedControl"] button[aria-pressed="true"] {
            background-color: #ffffff !important;
            color: #0f172a !important;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08) !important;
        }

        /* Premium Green Button */
        div[data-testid="stButton"] > button {
            background-color: #166534 !important;
            background: linear-gradient(180deg, #166534 0%, #14532b 100%) !important;
            color: #ffffff !important;
            border: none !important;
            border-radius: 8px !important;
            padding: 14px 16px !important;
            font-weight: 700 !important;
            font-size: 14.5px !important;
            width: 100% !important;
            box-shadow: 0 4px 14px rgba(22, 101, 52, 0.25) !important;
            transition: all 0.18s ease !important;
            margin-top: 8px !important;
        }
        div[data-testid="stButton"] > button:hover {
            background-color: #14532b !important;
            transform: translateY(-1px) !important;
            box-shadow: 0 6px 18px rgba(22, 101, 52, 0.32) !important;
        }
        div[data-testid="stButton"] > button:active {
            transform: translateY(0px) !important;
        }

        /* Mobile responsiveness */
        @media (max-width: 900px) {
            div[data-testid="stHorizontalBlock"] {
                flex-direction: column !important;
            }
            div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:first-child,
            div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:last-child {
                min-height: auto !important;
                padding: 40px 24px !important;
            }
        }
    </style>
    ''', unsafe_allow_html=True)

    col_left, col_right = st.columns([1, 1], gap=0)

    with col_left:
        st.markdown('''
        <div class="brand-container">
            <div class="brand-kicker">● GSTC · TGA UYUMLU PLATFORM</div>
            <div class="brand-title">Karbon<span>AT</span></div>
            <div class="brand-sub">Yapay Zeka Destekli Yeşil Dönüşüm</div>
            <div class="brand-desc">Oteller için otonom sürdürülebilirlik altyapısı — veriden rapora tek akışta.</div>
            <div class="bullet-item">
                <div class="bullet-icon">⚡</div>
                <div class="bullet-text">
                    <b>Otonom Veri Çıkarımı</b>
                    <span>Tüketim verilerini otomatik toplayın ve işleyin</span>
                </div>
            </div>
            <div class="bullet-item">
                <div class="bullet-icon">◈</div>
                <div class="bullet-text">
                    <b>Kapsam 1-2-3 Hesabı</b>
                    <span>HCMI metodolojisiyle eksiksiz karbon hesabı</span>
                </div>
            </div>
            <div class="bullet-item">
                <div class="bullet-icon">✓</div>
                <div class="bullet-text">
                    <b>TGA/GSTC Uyumlu Raporlama</b>
                    <span>Tablo 6-13, yeşil rapor ve denetime hazır çıktılar</span>
                </div>
            </div>
        </div>
        ''', unsafe_allow_html=True)

    with col_right:
        st.markdown('''
        <div class="auth-head-title">Hoş geldiniz</div>
        <div class="auth-head-sub">Hesabınıza giriş yapın veya yeni hesap oluşturun</div>
        ''', unsafe_allow_html=True)

        secim = st.segmented_control(
            "Giriş türü",
            options=["Giriş Yap", "Kayıt Ol"],
            default="Giriş Yap",
            key="auth_mode",
            label_visibility="hidden",
        )

        if secim == "Giriş Yap":
            username = st.text_input("Kullanıcı adı", placeholder="ornek_kullanici", key="auth_uname")
            password = st.text_input("Şifre", type="password", placeholder="••••••••", key="auth_pw")
            if st.button("Sisteme Giriş Yap", width='stretch'):
                user = verify_login(username, password)
                if user:
                    st.session_state.user = user
                    st.session_state.step = 0
                    st.rerun()
                else:
                    st.error("Kullanıcı adı veya şifre hatalı.")
        else:
            yeni_adi = st.text_input("Ad Soyad", placeholder="Ad Soyad", key="reg_name")
            yeni_uname = st.text_input("Kullanıcı adı", placeholder="kullanici_adi", key="reg_uname")
            yeni_pw = st.text_input("Şifre (en az 4 karakter)", type="password", placeholder="••••••••", key="reg_pw")
            if st.button("Kayıt Ol", width='stretch'):
                if len(yeni_pw) < 4:
                    st.error("Şifre en az 4 karakter olmalı.")
                else:
                    user = create_user(yeni_uname, yeni_pw, yeni_adi)
                    if user:
                        st.session_state.user = user
                        st.session_state.step = 0
                        st.rerun()
                    else:
                        st.error("Bu kullanıcı adı zaten alınmış.")
# ==============================================
# ADIM 0 - TESİS SEÇİMİ (PROFİLİM / OTELLERİM)
# ==============================================
def adim_tesis_secimi():
    st.markdown(
        f"""<div class="hero" style="padding:28px 18px 20px;">
            <div style="display:inline-flex; align-items:center; gap:8px; background:#fff; border:1px solid {BORDER}; padding:6px 10px; border-radius:999px; font-size:12px; font-weight:800; color:{PRIMARY}; box-shadow: var(--shadow-sm);">🌿 KarbonAT P2 · GSTC/TGA</div>
            <h1 style="margin:12px 0 6px; font-size:34px;">Otellerinizi yönetin</h1>
            <p style="max-width:640px; margin:0 auto; color:{MUTED};">Aylık tüketimi girin, TGA tablolarınız ve yeşil raporlarınız otomatik oluşsun.</p>
        </div>""",
        unsafe_allow_html=True,
    )
    user = st.session_state.user
    facilities = list_facilities(user["id"])
    if facilities:
        st.markdown(
            f'<div class="section-title">🏨 Profilim · Otellerim '
            f'<span style="color:{MUTED}; font-weight:700; background:#fff; border:1px solid {BORDER}; padding:3px 8px; border-radius:999px;">{len(facilities)}</span></div>',
            unsafe_allow_html=True,
        )
        col = st.columns([2, 3, 2])[1]
        with col:
            secenekler = {t["ad"]: t["id"] for t in facilities}
            secim = st.selectbox(
                "Tesisinizi seçin",
                list(secenekler.keys()),
                key="fac_select",
                on_change=_clear_facility_state,
            )
            new_fac_id = secenekler[secim]
            if st.session_state.facility_id != new_fac_id:
                _clear_facility_state()
                st.session_state.facility_id = new_fac_id
                st.session_state.tesis = get_facility(new_fac_id)
            else:
                st.session_state.facility_id = new_fac_id
                st.session_state.tesis = get_facility(new_fac_id)

            kayitlar = list_records(st.session_state.facility_id)
            if kayitlar:
                son = kayitlar[-1]
                st.markdown(
                    f'<div style="background: linear-gradient(180deg, #fff, #f6faf7); border:1px solid {BORDER}; border-radius:14px; padding:12px 14px; display:flex; align-items:center; justify-content:space-between; box-shadow: var(--shadow-sm);">'
                    f'<span style="display:flex; align-items:center; gap:8px;"><span style="width:28px; height:28px; border-radius:8px; background:{MIST}; border:1px solid {BORDER}; display:inline-flex; align-items:center; justify-content:center;">📅</span> <b>{format_donem(son["period"])}</b> · {son["sonuc"]["metrikler"]["toplam_ton"]} ton CO₂e</span>'
                    f'<span style="background:{PRIMARY}; color:#fff; padding:4px 10px; border-radius:999px; font-size:12px; font-weight:800;">✓ Kayıtlı</span></div>',
                    unsafe_allow_html=True,
                )

            st.markdown("")
            a, b = st.columns(2)
            with a:
                if st.button("📊 Raporlar", width='stretch'):
                    st.session_state.show_raporlar = True
                    st.rerun()
            with b:
                if st.button("📣 Medya & İçerik", width='stretch', type="secondary"):
                    st.session_state.step = 5
                    st.rerun()
            c1, c2, c3 = st.columns(3)
            with c1:
                if st.button("📊 Veri Girişi", width='stretch', type="secondary"):
                    st.session_state.history = list_records(st.session_state.facility_id)
                    st.session_state.step = 2
                    st.rerun()
            with c2:
                if st.button("✏️ Profili Düzenle", width='stretch', type="secondary"):
                    st.session_state.step = 1
                    st.rerun()
            with c3:
                if st.button("＋ Yeni Tesis", width='stretch', type="secondary"):
                    st.session_state.tesis = {}
                    st.session_state.facility_id = None
                    st.session_state.step = 1
                    st.rerun()

        st.markdown("---")
        with st.expander(
            "📊 Raporlar — aylara göre",
            expanded=bool(st.session_state.get("show_raporlar", False)),
        ):
            donemler = list_records(st.session_state.facility_id)
            if not donemler:
                st.caption(
                    "Henüz rapor yok. Önce veri girişi + hesaplama yapın; sonuç ekranındaki "
                    "📊 Raporlar bölümü raporlarınızı üretir ve burada aylara göre saklanır."
                )
            else:
                secenek = {format_donem(r["period"]): r["period"] for r in donemler}
                sec_period = st.selectbox("Dönem seçin", list(secenek), key="profil_rapor_period")
                rec = get_record(st.session_state.facility_id, secenek[sec_period])
                if rec:
                    _rapor_karti(st.session_state.facility_id, rec["period"], rec["sonuc"])
    else:
        col = st.columns([2, 3, 2])[1]
        with col:
            st.markdown(
                f'<div style="background: linear-gradient(180deg, #fff, #f7faf7); border:1px solid {BORDER}; border-radius:18px; padding:22px 18px; text-align:center; box-shadow: var(--shadow-md);">'
                f'<div style="width:56px; height:56px; border-radius:14px; background: linear-gradient(135deg,{PRIMARY},{ACCENT}); color:#fff; display:flex; align-items:center; justify-content:center; font-size:26px; margin:0 auto 10px; box-shadow:0 8px 16px rgba(29,107,69,.22);">🏨</div>'
                f'<div style="font-weight:800; font-size:16px;">Henüz oteliniz yok</div>'
                f'<div style="font-size:13px; color:{MUTED}; margin:6px 0 14px;">Hoş geldiniz, <b>{user.get("fullname", user["username"])}</b>! İlk tesis profilinizi oluşturarak başlayın — 1 dakikada hazır.</div></div>',
                unsafe_allow_html=True,
            )
            st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)
            if st.button("▶ İlk Tesisi Oluştur", width='stretch'):
                st.session_state.step = 1
                st.rerun()


# ==============================================
# ADIM 1 - TESİS PROFİLİ
# ==============================================
def adim_tesis():
    st.markdown(mvp_stepper(active=0), unsafe_allow_html=True)
    st.markdown('<h1>🏨 Tesis Profili</h1>', unsafe_allow_html=True)
    st.markdown('<p style="color:#6b7a70;">Bir kerelik bilgiler. Sonradan değiştirilebilir. <b>İpucu:</b> m² = ısıtılan toplam alan, oda = envanterdeki oda sayısı.</p>', unsafe_allow_html=True)

    t = st.session_state.tesis or {}
    st.markdown('<div class="section-title">Kimlik & Kapasite</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        tesis_adi = st.text_input("Tesis / Otel adı", value=t.get("ad", ""), placeholder="Örn. Grand Horizon")
    with col2:
        m2 = st.number_input("Toplam kapalı alan (m²)", min_value=1, value=int(t.get("m2", 1000)), step=100)
    with col3:
        oda = st.number_input("Toplam oda sayısı", min_value=1, value=int(t.get("oda", 50)), step=10)

    col4, col5 = st.columns(2)
    with col4:
        personel = st.number_input("Personel sayısı", min_value=1, value=int(t.get("personel", 20)), step=5)
    with col5:
        st.caption("Müşteri, oda-gün, atık bertarafı ve yenilenebilir oranı aylık veri girişinde girilir.")

    st.markdown("---")
    col_back, col_next = st.columns([1, 2])
    with col_back:
        if st.button("← Geri", width='stretch'):
            st.session_state.step = 0
            st.rerun()
    with col_next:
        disabled = not (tesis_adi and m2 and oda and personel)
        if st.button("💾 Kaydet ve Devam", width='stretch', disabled=disabled):
            tesis = {
                "ad": tesis_adi,
                "m2": int(m2),
                "oda": int(oda),
                "personel": int(personel),
            }
            if st.session_state.facility_id:
                tesis["id"] = st.session_state.facility_id
            tesis = save_facility(tesis, owner_id=st.session_state.user["id"])
            st.session_state.tesis = tesis
            st.session_state.facility_id = tesis["id"]
            st.session_state.history = list_records(tesis["id"])
            st.session_state.step = 2
            st.rerun()


# ==============================================
# ADIM 2 - AYLIK VERİ (KART BAZLI GİRİŞ)
# ==============================================
# A1 sadeleştirme: aylık özel alanlar (atik bertaraf, yenilenebilir oranı)
ATIK_BERTARAF_SECENEKLERI = [
    "Geri dönüşüm + çöp (karışık)",
    "Ağırlıklı geri dönüşüm",
    "Çoğunlukla çöp + kompost",
]


def _eski_bertaraf(tesis, mevcut, onceki):
    """Backward compat: tesis dict'inden ya da eski kayıttan."""
    if mevcut and mevcut.get("atik_bertaraf"):
        return mevcut["atik_bertaraf"]
    if mevcut and mevcut.get("atik_bertaraf_idx") is not None:
        idx = mevcut["atik_bertaraf_idx"]
        if 0 <= idx < len(ATIK_BERTARAF_SECENEKLERI):
            return ATIK_BERTARAF_SECENEKLERI[idx]
    if tesis.get("atik_bertaraf"):
        return tesis["atik_bertaraf"]
    if onceki and onceki.get("atik_bertaraf"):
        return onceki["atik_bertaraf"]
    return ATIK_BERTARAF_SECENEKLERI[0]


def _eski_yenilenebilir(tesis, mevcut, onceki):
    for src in (mevcut, onceki, tesis):
        if src and src.get("yenilenebilir_oran") is not None:
            return int(src["yenilenebilir_oran"])
        if src and src.get("yenilenebilir") is not None:
            return int(src["yenilenebilir"])
    return 30


def _demo_veri_uret(tesis):
    """Tesis kapasitesine göre gerçekçi demo tüketim üretir (görüşmede 1 tık değer görme)."""
    import random
    m2 = int(tesis.get("m2", 1000))
    oda = int(tesis.get("oda", 50))
    pers = int(tesis.get("personel", 20))
    # Operasyon: musteri ~ oda*0.65*30, oda-gün ~ oda*0.62*30
    musteri = max(120, int(oda * 0.65 * 30 * random.uniform(0.85, 1.05)))
    dolu_oda_gun = max(300, int(oda * 0.62 * 30 * random.uniform(0.85, 1.05)))
    # Elektrik kWh ~ m2*14 + oda*55
    elec_total = m2 * random.uniform(10, 15) + oda * 55 + pers * 30
    tuketim = {}
    for kat in EMISSION_FACTORS:
        tuketim[kat] = {alt: 0.0 for alt in EMISSION_FACTORS[kat]}
    # Elektrik dağılımı
    tuketim["Elektrik"]["Şebeke (yenilenebilir olmayan)"] = round(elec_total * 0.68, 1)
    tuketim["Elektrik"]["Şebeke (yenilenebilir YEK-G sertifikalı)"] = round(elec_total * 0.18, 1)
    tuketim["Elektrik"]["Güneş Enerjisi (PV)"] = round(elec_total * 0.09, 1)
    tuketim["Elektrik"]["Rüzgar Enerjisi"] = round(elec_total * 0.05, 1)
    # Doğal gaz & yakıt
    tuketim["Doğal Gaz ve Yakıtlar"]["Doğalgaz (m³)"] = round(m2 * 0.7 * random.uniform(0.8, 1.1), 1)
    tuketim["Doğal Gaz ve Yakıtlar"]["LPG (kg)"] = round(oda * 1.2 * random.uniform(0.7, 1.2), 1)
    # F-gaz kaçak
    tuketim["Soğutucu & F-Gaz (Scope 1)"]["R-410A (kg)"] = round(random.uniform(1.5, 4.5), 2)
    tuketim["Soğutucu & F-Gaz (Scope 1)"]["R-32 (kg)"] = round(random.uniform(0.5, 2.0), 2)
    # Araç
    tuketim["Araç Filosu & İş Seyahatleri (Scope 1/3)"]["Benzinli Araç (L)"] = round(180 + oda * 1.5, 1)
    tuketim["Araç Filosu & İş Seyahatleri (Scope 1/3)"]["Dizel Araç (L)"] = round(220 + oda * 1.2, 1)
    # Su
    tuketim["Su"]["Şebeke suyu tüketimi (m³)"] = round(musteri * 0.18 + oda * 0.9, 1)
    tuketim["Su"]["Atık su arıtma (m³)"] = round(musteri * 0.12, 1)
    # Gıda
    tuketim["Gıda Tüketimi"]["Kırmızı Et (kg)"] = round(musteri * 0.25, 1)
    tuketim["Gıda Tüketimi"]["Tavuk (kg)"] = round(musteri * 0.35, 1)
    tuketim["Gıda Tüketimi"]["Sebze (kg)"] = round(musteri * 0.9, 1)
    tuketim["Gıda Tüketimi"]["Süt (kg)"] = round(musteri * 0.4, 1)
    # Atık
    tuketim["Atık Yönetimi"]["Organik Atık (kg)"] = round(musteri * 0.55, 1)
    tuketim["Atık Yönetimi"]["Plastik Atık (kg)"] = round(musteri * 0.18, 1)
    tuketim["Atık Yönetimi"]["Kağıt (kg)"] = round(musteri * 0.12, 1)
    # Kimyasal
    tuketim["Kimyasal Tüketimi"]["Deterjan (L)"] = round(oda * 0.6, 1)
    tuketim["Kimyasal Tüketimi"]["Temizlik Ürünleri (L)"] = round(oda * 0.9, 1)
    return {
        "tuketim": tuketim,
        "musteri": musteri,
        "dolu_oda_gun": dolu_oda_gun,
        "atik_bertaraf": "Ağırlıklı geri dönüşüm",
        "yenilenebilir_oran": random.choice([22, 28, 35]),
    }


def _kategori_kart(kategori, tuketim, period):
    """Bir kategori için alt tür kartları (2 kolon)."""
    alt_turler = EMISSION_FACTORS.get(kategori, {})
    birim = KATEGORI_BIRIM.get(kategori, "")
    cols = st.columns(2)
    yeni = {}
    for i, (ad, ef) in enumerate(alt_turler.items()):
        col = cols[i % 2]
        with col:
            st.markdown(
                f'<div class="tga-card">'
                f'<div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:10px;">'
                f'<div style="display:flex; align-items:center; gap:8px;">'
                f'<span style="font-weight:700; font-size:14px; color:#0f172a; letter-spacing:-0.2px;">{ad}</span>'
                f'<span style="background:#f1f5f9; color:#475569; font-weight:600; font-size:11px; padding:2px 8px; border-radius:4px;">{birim}</span>'
                f'</div>'
                f'<span style="background:#f8fafc; color:#64748b; border:1px solid #e2e8f0; font-weight:600; font-size:11px; padding:2px 8px; border-radius:4px;">EF {ef} kgCO₂e/{birim}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            mevcut_deger = float(tuketim.get(kategori, {}).get(ad, 0.0))
            yeni_deger = st.number_input(
                "Miktar",
                min_value=0.0,
                step=1.0,
                value=mevcut_deger,
                key=f"veri_{st.session_state.facility_id}_{period}_{kategori}_{ad}",
                label_visibility="hidden",
                format="%.2f",
                placeholder="0.00",
            )
            yeni[ad] = yeni_deger
            st.markdown('</div>', unsafe_allow_html=True)
    return yeni


def adim_veri():
    st.markdown(mvp_stepper(active=1), unsafe_allow_html=True)
    tesis = st.session_state.tesis
    st.markdown(
        f"""<div style="display:flex; align-items:center; gap:12px; margin:6px 0 4px;">
            <div style="width:42px;height:42px;border-radius:12px; background: linear-gradient(135deg,{PRIMARY},{ACCENT}); display:flex; align-items:center; justify-content:center; color:#fff; font-size:20px; box-shadow:0 6px 16px rgba(29,107,69,.22);">📊</div>
            <div><div style="font-size:22px; font-weight:800; letter-spacing:-0.5px; line-height:1;">Aylık Veri Girişi</div>
            <div style="font-size:13px; color:{MUTED}; margin-top:2px;">TGA takip tabloları için bu ayın tüketim verileri</div></div></div>""",
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<div style="display:flex; flex-wrap:wrap; gap:8px; margin:10px 0 2px;">'
        f'<span class="chip">🌿 {tesis["ad"]}</span>'
        f'<span class="chip">📐 {_tesis_ozet(tesis)}</span>'
        f'<span class="chip" style="background:#fff8ec; border-color:#f0d9a0; color:#8a5a12;">🧾 {st.session_state.period}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # Dönem seçimi
    options = _period_options()
    current = st.session_state.period
    if current not in options:
        options.insert(0, current)
    idx = options.index(current) if current in options else 0
    secilen = st.selectbox(
        "📅 Veri dönemi",
        options,
        index=idx,
        format_func=format_donem,
    )
    st.session_state.period = secilen

    # Kayıt / önceki dönem yükle
    mevcut = get_record(st.session_state.facility_id, secilen)
    onceki = get_previous_record(st.session_state.facility_id, secilen)

    if mevcut:
        st.markdown(
            f'<div class="banner">✅ {format_donem(secilen)} için kayıt var. '
            f'Değerleri düzenleyip yeniden hesaplayabilirsiniz.</div>',
            unsafe_allow_html=True,
        )
    elif onceki:
        st.markdown(
            f'<div class="banner">📎 Önceki dönem ({format_donem(onceki["period"])}) '
            f'değerleri alt türlere yüklendi. Düzenleyebilirsiniz.</div>',
            unsafe_allow_html=True,
        )

    # Tuketim + aylık alanlar (tesis+donem bazlı anahtar)
    _load_key = f"{st.session_state.facility_id}::{secilen}"
    if "_load_period" not in st.session_state or st.session_state._load_period != _load_key:
        base = None
        if mevcut:
            base = mevcut["tuketim"]
        elif onceki:
            base = onceki["tuketim"]
        tuk = {}
        for kat in EMISSION_FACTORS:
            tuk[kat] = {}
            for alt in EMISSION_FACTORS[kat]:
                tuk[kat][alt] = base.get(kat, {}).get(alt, 0.0) if base else 0.0
        st.session_state.tuketim = tuk
        st.session_state.musteri = mevcut["musteri"] if mevcut else (onceki["musteri"] if onceki else 0)
        st.session_state.dolu_oda_gun = mevcut["dolu_oda_gun"] if mevcut else (onceki["dolu_oda_gun"] if onceki else 0)
        st.session_state.atik_bertaraf = _eski_bertaraf(tesis, mevcut, onceki)
        st.session_state.yenilenebilir_oran = _eski_yenilenebilir(tesis, mevcut, onceki)
        st.session_state._load_period = _load_key

    tuketim = st.session_state.tuketim

    # ——— 1-tık demo: tesis kapasitesine göre gerçekçi örnek veri ———
    _toplam_girdi = sum(v for kat in tuketim.values() for v in kat.values()) + int(st.session_state.musteri) + int(st.session_state.dolu_oda_gun)
    if _toplam_girdi == 0:
        st.markdown(
            f'<div class="mvp-empty" style="text-align:left; display:flex; gap:14px; align-items:center;">'
            f'<div style="width:44px; height:44px; border-radius:12px; background: linear-gradient(135deg,{PRIMARY},{ACCENT}); color:#fff; display:flex; align-items:center; justify-content:center; font-size:20px; flex-shrink:0;">✨</div>'
            f'<div style="flex:1;"><div style="font-weight:800;">İlk değerinizi hemen görün</div>'
            f'<div style="font-size:13px; color:{MUTED};">{tesis["ad"]} kapasitesine göre örnek tüketim oluşturayım mı? Tek tıkla doldurup hemen hesaplayabilirsiniz — sonra istediğinizi düzenlersiniz.</div></div></div>',
            unsafe_allow_html=True,
        )
        if st.button("✨ Kapasiteye göre örnek veriyle doldur — 1 tık", width='stretch'):
            demo = _demo_veri_uret(tesis)
            # session + widget state'i doğrudan demo değerleriyle doldur
            st.session_state.tuketim = demo["tuketim"]
            st.session_state.musteri = demo["musteri"]
            st.session_state.dolu_oda_gun = demo["dolu_oda_gun"]
            st.session_state.atik_bertaraf = demo["atik_bertaraf"]
            st.session_state.yenilenebilir_oran = demo["yenilenebilir_oran"]
            for kat, alts in demo["tuketim"].items():
                for alt, val in alts.items():
                    wk = f"veri_{st.session_state.facility_id}_{secilen}_{kat}_{alt}"
                    st.session_state[wk] = float(val)
            # hemen hesaplayıp kaydet ki Sonuç ekranı 0 göstermesin
            try:
                _demo_sonuc = _hesapla(tesis, demo["tuketim"], demo["musteri"], demo["dolu_oda_gun"],
                                       atik_bertaraf=demo["atik_bertaraf"], yenilenebilir_oran=demo["yenilenebilir_oran"])
                save_record({
                    "fac_id": st.session_state.facility_id,
                    "period": secilen,
                    "musteri": demo["musteri"],
                    "dolu_oda_gun": demo["dolu_oda_gun"],
                    "tuketim": demo["tuketim"],
                    "atik_bertaraf": demo["atik_bertaraf"],
                    "yenilenebilir_oran": demo["yenilenebilir_oran"],
                    "sonuc": _demo_sonuc,
                })
                st.session_state.sonuc = _demo_sonuc
                st.session_state.history = list_records(st.session_state.facility_id)
            except Exception:
                pass
            st.toast("Örnek veri dolduruldu — Hesaplandı, Sonuç ekranına geçebilirsiniz", icon="✨")
            st.rerun()
            st.stop()

    # Operasyon
    st.markdown('<div class="section-title">🛏️ Bu Ayki Operasyon</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:12px;">', unsafe_allow_html=True)
    col_m, col_d = st.columns(2)
    with col_m:
        st.markdown(f'<div class="data-card" style="margin:0;">'
                    f'<div style="font-size:11px; font-weight:800; letter-spacing:.08em; text-transform:uppercase; color:{MUTED}; margin-bottom:6px;">👥 Müşteri</div>', unsafe_allow_html=True)
        musteri = st.number_input(
            "Müşteri sayısı (konaklayan kişi)", min_value=0, step=10,
            value=int(st.session_state.musteri), help="Örn: 1 aile = 4 kişi",
            label_visibility="collapsed",
        )
        st.markdown('</div>', unsafe_allow_html=True)
    with col_d:
        st.markdown(f'<div class="data-card" style="margin:0;">'
                    f'<div style="font-size:11px; font-weight:800; letter-spacing:.08em; text-transform:uppercase; color:{MUTED}; margin-bottom:6px;">🛏️ Satılan Oda-Gün</div>', unsafe_allow_html=True)
        dolu_oda_gun = st.number_input(
            "Satılan oda-gün sayısı", min_value=0, step=50,
            value=int(st.session_state.dolu_oda_gun),
            help="Toplam oda × doluluk × gün (HCMI)",
            label_visibility="collapsed",
        )
        st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.session_state.musteri = musteri
    st.session_state.dolu_oda_gun = dolu_oda_gun

    # TGA tablolarına göre sekmeler
    st.markdown('<div class="section-title">🗂️ TGA Takip Tabloları</div>', unsafe_allow_html=True)

    sekmeler = [
        ("Elektrik (Tablo 10)", "Elektrik"),
        ("Doğal Gaz ve Yakıtlar", "Doğal Gaz ve Yakıtlar"),
        ("Soğutucu & F-Gaz", "Soğutucu & F-Gaz (Scope 1)"),
        ("Araç Filosu & Seyahat", "Araç Filosu & İş Seyahatleri (Scope 1/3)"),
        ("Su (Tablo 12)", "Su"),
        ("Gıda Tüketimi", "Gıda Tüketimi"),
        ("Atık Yönetimi (Tablo 13)", "Atık Yönetimi"),
        ("Kimyasal Tüketimi", "Kimyasal Tüketimi"),
    ]

    tabs = st.tabs([s[0] for s in sekmeler])
    for tab, (_, kat) in zip(tabs, sekmeler):
        with tab:
            aciklama = KATEGORI_ACIKLAMALARI.get(kat, "")
            if aciklama:
                st.caption(aciklama)

            # Elektrik sekmesi: yenilenebilir oranı (opsiyonel)
            if kat == "Elektrik":
                st.markdown(
                    f'<div class="data-card" style="padding:12px 14px; margin:6px 0;">',
                    unsafe_allow_html=True,
                )
                st.session_state.yenilenebilir_oran = st.slider(
                    "Yenilenebilir elektrik oranı (%)", 0, 100,
                    value=int(st.session_state.yenilenebilir_oran),
                    step=5,
                    help="Şebeke elektriğinin YEK-G sertifikalı (yenilenebilir) kısmı. "
                         "Boş bırakmak isterseniz 0 ayarlayın ya da alt türlerde "
                         "'Şebeke (yenilenebilir YEK-G sertifikalı)' satırını kullanın.",
                )
                st.markdown('</div>', unsafe_allow_html=True)

            # Atık sekmesi: ağırlıklı bertaraf yöntemi
            if kat == "Atık Yönetimi":
                st.markdown(
                    f'<div class="data-card" style="padding:12px 14px; margin:6px 0;">',
                    unsafe_allow_html=True,
                )
                idx_bertaraf = (
                    ATIK_BERTARAF_SECENEKLERI.index(st.session_state.atik_bertaraf)
                    if st.session_state.atik_bertaraf in ATIK_BERTARAF_SECENEKLERI else 0
                )
                st.session_state.atik_bertaraf = st.selectbox(
                    "Ağırlıklı atık bertaraf yöntemi",
                    ATIK_BERTARAF_SECENEKLERI,
                    index=idx_bertaraf,
                )
                st.markdown('</div>', unsafe_allow_html=True)

            tuketim[kat] = _kategori_kart(kat, tuketim, secilen)

            kat_toplam = sum(
                miktar * EMISSION_FACTORS[kat].get(alt, 0.0)
                for alt, miktar in tuketim[kat].items()
            )
            aktif = sum(1 for v in tuketim[kat].values() if v > 0)
            c1, c2 = st.columns([1, 2])
            with c1:
                st.metric("Kategori Emisyonu", f"{kat_toplam:,.1f} kg CO₂e")
            with c2:
                st.caption(f"{aktif} alt tür dolu · Emisyon otomatik hesaplanır")

    st.session_state.tuketim = tuketim

    # Canlı özet
    st.markdown('<div class="section-title">🧮 Canlı Özet</div>', unsafe_allow_html=True)
    onizleme = _hesapla(
        tesis, tuketim, musteri, dolu_oda_gun,
        atik_bertaraf=st.session_state.atik_bertaraf,
        yenilenebilir_oran=st.session_state.yenilenebilir_oran,
    )
    m = onizleme["metrikler"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Toplam", f"{m['toplam_ton']} ton CO₂e")
    c2.metric("Oda-Gün", f"{m['oda_gun_kg']} kg")
    c3.metric("Müşteri Başına", f"{m['musteri_kg']} kg")
    c4.metric("m² Başına", f"{m['m2_aylik_kg']} kg")

    st.markdown('<div class="mvp-cta">', unsafe_allow_html=True)
    st.markdown("---")
    col_back, col_draft, col_prev, col_next = st.columns([1, 1, 1, 2])
    with col_back:
        if st.button("← Geri", width='stretch'):
            st.session_state.step = 1
            st.rerun()
    with col_draft:
        if st.button("💾 Taslak Kaydet", width='stretch'):
            _save_draft(tesis["id"])
            st.toast("Taslak kaydedildi", icon="💾")
    with col_prev:
        if st.button("🔙 Önceki Aya Dön", width='stretch'):
            _load_previous_period()
            st.rerun()
    with col_next:
        if st.button("🧮 Hesapla ve Raporla", width='stretch'):
            st.session_state.step = 3
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ==============================================
# ADIM 3 - HESAPLAMA
# ==============================================
def adim_hesap():
    st.markdown(mvp_stepper(active=2), unsafe_allow_html=True)
    st.markdown('<h1>🧮 Hesaplama</h1>', unsafe_allow_html=True)
    tesis = st.session_state.tesis
    period = st.session_state.period
    tuketim = st.session_state.tuketim

    with st.spinner("Hesaplanıyor..."):
        sonuc = _hesapla(
            tesis, tuketim,
            st.session_state.musteri,
            st.session_state.dolu_oda_gun,
            atik_bertaraf=st.session_state.atik_bertaraf,
            yenilenebilir_oran=st.session_state.yenilenebilir_oran,
        )
        save_record({
            "fac_id": st.session_state.facility_id,
            "period": period,
            "musteri": st.session_state.musteri,
            "dolu_oda_gun": st.session_state.dolu_oda_gun,
            "tuketim": tuketim,
            "atik_bertaraf": st.session_state.atik_bertaraf,
            "atik_bertaraf_idx": ATIK_BERTARAF_SECENEKLERI.index(st.session_state.atik_bertaraf),
            "yenilenebilir_oran": st.session_state.yenilenebilir_oran,
            "sonuc": sonuc,
        })
        st.session_state.history = list_records(st.session_state.facility_id)

    st.session_state.sonuc = sonuc
    st.session_state.step = 4
    st.rerun()


# ==============================================
# ADIM 4 - SONUÇ + RAPOR (TEKİLLEŞTİRİLMİŞ)
# ==============================================
# RAPORLAR (raporlama alt sistemi) — sonuç + profil'de paylaşılır
# ==============================================
def _markdown_blok(metin):
    """Markdown'ı gösterir; tabloları kaydırmalı DataFrame olarak basar (taşma olmaz)."""
    for tur, icerik in raporlar.markdown_bloklar(metin):
        if tur == "tablo":
            st.dataframe(icerik, width='stretch', hide_index=True)
        elif icerik and icerik.strip():
            st.markdown(icerik)


def _rapor_detay(fac_id, period, sab, sonuc, tesis, prefs):
    sab_id = sab["id"]
    rte_key = f"rte_{fac_id}_{period}_{sab_id}"
    mod_key = f"rmod_{fac_id}_{period}_{sab_id}"

    h1, h2 = st.columns([6, 1])
    with h1:
        st.markdown(f"### {sab['emoji']} {sab['baslik']}")
    with h2:
        if st.button("✖ Kapat", key=f"rkapat_{fac_id}_{period}_{sab_id}"):
            st.session_state.pop(f"rapor_detay_{fac_id}_{period}", None)
            st.rerun()
    st.caption(sab["aciklama"])

    kayit = get_report(fac_id, period, sab_id)

    if st.button("🚀 Üret / Yenile", key=f"rug_{fac_id}_{period}_{sab_id}", width='stretch'):
        try:
            with st.spinner("Şablon tesis verilerinizle dolduruluyor..."):
                cik = raporlar.rapor_uretim(sab_id, tesis, sonuc, prefs)
            save_report(fac_id, period, sab_id, cik["metin"], tip=cik["tip"])
            st.session_state.pop(rte_key, None)
            st.toast("Rapor kaydedildi; profilinizde aylara göre görebilirsiniz.")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Üretim başarısız: {e}")
            sy = str(e)
            if "429" in sy or "RESOURCE_EXHAUSTED" in sy or "quota" in sy.lower():
                st.warning(
                    "Kota (429) aşıldı — ücretsiz katmanın günlük/dakikalık limiti "
                    "tükendi. Dakika kotası birkaç saniyede açılır; üst üste deneme "
                    "engellemeyi uzatır. Kalıcı çözüm: GEMINI_API_KEY'i ücretli bir "
                    "anahtarla değiştirin. https://ai.google.dev/gemini-api/docs/rate-limits"
                )
                # Fallback: boş ekran yerine deterministik taslak kaydet
                try:
                    fallback = _ai_fallback_rapor(sab, tesis, sonuc)
                    save_report(fac_id, period, sab_id, fallback, tip="ai-fallback")
                    st.session_state.pop(rte_key, None)
                    st.info("AI kotası dolu — geçici deterministik taslak kaydedildi. Aşağıda açıldı; birkaç dakika sonra Yenile ile zenginleştirebilirsiniz.")
                    st.rerun()
                except Exception:
                    pass
            else:
                st.caption("İpucu: GEMINI_API_KEY .env içinde olmalı; kota yoksa birkaç saniye sonra tekrar deneyin.")

    if not kayit:
        st.info("Henüz üretilmedi — yukarıdaki düğmeyle üretin; içerik düzenlenebilir ve PDF/Word/Excel indirilebilir.")
        return

    icerik = st.session_state.get(rte_key) or kayit["metin"]

    st.markdown("### ✍️ İçerik")
    mod = st.segmented_control(
        "Görünüm",
        options=["✏️ Düzenle", "👁 Önizleme"],
        default="✏️ Düzenle",
        key=mod_key,
        label_visibility="hidden",
    )

    if mod == "✏️ Düzenle":
        icerik = st.text_area(
            "Rapor içeriği — markdown yazabilirsiniz",
            value=icerik,
            height=520,
            key=rte_key,
        )
        st.caption("Başlık: `### Başlık`  ·  Madde: `- metin`  ·  Tablo: `| Sütun | Değer |` satırları.")
    else:
        st.markdown("---")
        _markdown_blok(icerik)
        st.caption("Tablolar kaydırmalı tablo olarak gösterilir; PDF/Word/Excel gerçek tablo içerir.")

    st.markdown("---")
    d1, d2, d3, d4 = st.columns(4)
    with d1:
        try:
            st.download_button("⬇️ PDF", data=raporlar.rapor_pdf(icerik),
                               file_name=f"KarbonAT_{sab_id}_{period}.pdf", mime="application/pdf",
                               key=f"rapdf_{fac_id}_{period}_{sab_id}", width='stretch')
        except Exception as e:  # noqa: BLE001
            st.warning(f"PDF: {e}")
    with d2:
        try:
            st.download_button("📄 Word (.docx)", data=raporlar.rapor_docx(icerik),
                               file_name=f"KarbonAT_{sab_id}_{period}.docx",
                               mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                               key=f"radoc_{fac_id}_{period}_{sab_id}", width='stretch')
        except Exception as e:  # noqa: BLE001
            st.warning(f"Word: {e}")
    with d3:
        try:
            if sab["tip"] == "deterministik":
                xlsx = raporlar.rapor_uretim(sab_id, tesis, sonuc)["xlsx"]
            else:
                xlsx = raporlar.rapor_xlsx(icerik)
            st.download_button("📊 Excel", data=xlsx,
                               file_name=f"KarbonAT_{sab_id}_{period}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"raex_{fac_id}_{period}_{sab_id}", width='stretch')
        except Exception as e:  # noqa: BLE001
            st.warning(f"Excel: {e}")
    with d4:
        if st.button("💾 Düzenlemeyi Kaydet", key=f"rasv_{fac_id}_{period}_{sab_id}", width='stretch'):
            save_report(fac_id, period, sab_id, icerik, tip=kayit.get("tip", "ai"))
            st.toast("Düzenleme kaydedildi.")
            st.rerun()


def _rapor_karti(fac_id, period, sonuc):
    if not sonuc:
        st.markdown(
            f'<div class="banner" style="border-left-color:{AMBER};">'
            f'<strong>⏳ Raporlar için hesaplama gerekiyor.</strong> Önce verinizi hesaplayın, sonra her şablon canlı veriyle dolar.</div>',
            unsafe_allow_html=True,
        )
        return
    tesis = sonuc.get("tesis") or {}
    st.markdown('<div class="section-title">📊 TGA Rapor Şablonları</div>', unsafe_allow_html=True)
    st.caption("Bir kart seçin → AI ile tesis verinizle doldurun → düzenleyin → PDF / Word / Excel indirin. Üretilenler profilinizde saklanır.")
    prefs = {"amac": "Raporlama", "ton": "Kurumsal & Resmi", "dil": "Türkçe", "uzunluk": "Detaylı"}
    secim_key = f"rapor_detay_{fac_id}_{period}"
    SUTUN = 3
    cols = st.columns(SUTUN)
    for i, sab in enumerate(raporlar.RAPOR_SABLONLARI):
        sab_id = sab["id"]
        kayit = get_report(fac_id, period, sab_id)
        ready = bool(kayit)
        badge = f'<span style="background:{"#e6f4ea" if ready else "#fff7e6"}; color:{"#0f5132" if ready else "#7a4a00"}; border:1px solid {"#b7e1c3" if ready else "#f0d9a0"}; padding:3px 8px; border-radius:999px; font-size:11px; font-weight:800;">{"✅ Hazır" if ready else "⏳ Beklemede"}</span>'
        desc = " · ".join(sab.get("cikti", []))
        with cols[i % SUTUN]:
            st.markdown(
                f'<div class="data-card" style="padding:14px 14px 10px; min-height:118px; display:flex; flex-direction:column; justify-content:space-between;">'
                f'<div><div style="font-size:18px; margin-bottom:6px;">{sab["emoji"]}</div>'
                f'<div style="font-weight:800; font-size:14px; line-height:1.25;">{sab["baslik"]}</div>'
                f'<div style="font-size:12px; color:{MUTED}; margin-top:4px; line-height:1.4;">{sab.get("aciklama","")[:92]}</div></div>'
                f'<div style="margin-top:10px; display:flex; align-items:center; justify-content:space-between; gap:8px;">{badge}<span style="font-size:11px; color:{MUTED};">{desc}</span></div></div>',
                unsafe_allow_html=True,
            )
            if st.button(f"→ Aç", key=f"rcd_{fac_id}_{period}_{sab_id}", width='stretch'):
                st.session_state[secim_key] = sab_id
                st.rerun()

    secilen = st.session_state.get(secim_key)
    if secilen:
        sab = raporlar.sablon_bul(secilen)
        if sab:
            st.markdown("---")
            _rapor_detay(fac_id, period, sab, sonuc, tesis, prefs)


# ==============================================
def adim_sonuc():
    st.markdown(mvp_stepper(active=3), unsafe_allow_html=True)
    st.markdown('<h1>📄 Karbon Ayak İzi Sonucu</h1>', unsafe_allow_html=True)

    if st.session_state.sonuc is None:
        st.warning("Önce hesaplama yapın.")
        return

    r = st.session_state.sonuc
    tesis = r["tesis"]
    scope = r["scope"]
    metrik = r["metrikler"]
    en_agir = r["en_agir"]
    period = st.session_state.period
    tuketim = r["tuketim"]

    st.markdown(
        f'<div class="chip">🌿 {tesis["ad"]}</div>'
        f'<div class="chip">📅 {format_donem(period)}</div>',
        unsafe_allow_html=True,
    )

    # KPI
    st.markdown('<div class="section-title">🔢 Temel Göstergeler</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-grid">
      <div class="kpi-card"><div class="kpi-label">🌍 Toplam Emisyon</div>
        <div class="kpi-value">{metrik['toplam_ton']} ton CO₂e</div>
        <div class="kpi-sub">{metrik['toplam_kg']:,.0f} kg · aylık</div></div>
      <div class="kpi-card"><div class="kpi-label">🛏️ Oda-Gün Başına</div>
        <div class="kpi-value">{metrik['oda_gun_kg']} kg</div>
        <div class="kpi-sub">HCMI standart metrik</div></div>
      <div class="kpi-card"><div class="kpi-label">📐 m² Başına</div>
        <div class="kpi-value">{metrik['m2_aylik_kg']} kg</div>
        <div class="kpi-sub">Aylık, m² başına</div></div>
      <div class="kpi-card"><div class="kpi-label">👤 Müşteri Başına</div>
        <div class="kpi-value">{metrik['musteri_kg']} kg</div>
        <div class="kpi-sub">Konaklayan kişi başına</div></div>
    </div>
    """, unsafe_allow_html=True)

    # Geçmiş karşılaştırma
    onceki = get_previous_record(st.session_state.facility_id, period)
    if onceki:
        om = onceki["sonuc"]["metrikler"]
        if om.get("toplam_kg", 0) > 0:
            delta = (metrik["toplam_kg"] - om["toplam_kg"]) / om["toplam_kg"] * 100
            yon = "azalma" if delta < 0 else "artış"
            renk = PRIMARY if delta < 0 else TERRA
            st.markdown(
                f'<div class="banner">📈 {format_donem(onceki["period"])} dönemine göre toplamda '
                f'<strong style="color:{renk};">%{abs(delta):.1f} {yon}</strong> '
                f'({om["toplam_kg"]:,.0f} kg → {metrik["toplam_kg"]:,.0f} kg)</div>',
                unsafe_allow_html=True,
            )

    col_left, col_right = st.columns([1, 1])

    # Scope dağılımı
    with col_left:
        st.markdown('<div class="section-title">🎯 Scope Dağılımı</div>', unsafe_allow_html=True)
        scope_toplam = metrik["scope1_kg"] + metrik["scope2_kg"] + metrik["scope3_kg"]
        if scope_toplam > 0:
            s1 = metrik["scope1_kg"] / scope_toplam * 100
            s2 = metrik["scope2_kg"] / scope_toplam * 100
            s3 = metrik["scope3_kg"] / scope_toplam * 100
        else:
            s1 = s2 = s3 = 0

        for baslik, deger, pct, renk in [
            ("Scope 1 · Doğrudan Yakıtlar", metrik["scope1_kg"], s1, PRIMARY),
            ("Scope 2 · Elektrik", metrik["scope2_kg"], s2, PRIMARY2),
            ("Scope 3 · Gıda, Su, Atık, Kimyasal", metrik["scope3_kg"], s3, ACCENT),
        ]:
            st.markdown(f"""
            <div class="data-card" style="padding:14px 16px; margin:8px 0;">
              <div style="display:flex; justify-content:space-between; margin-bottom:8px;">
                <span style="font-size:13.5px;"><strong>{baslik}</strong></span>
                <span style="font-weight:800; color:{PRIMARY};">{deger:,.0f} kg</span>
              </div>
              <div style="background:{BORDER}; height:10px; border-radius:6px; overflow:hidden;">
                <div style="background:{renk}; width:{pct:.1f}%; height:100%;"></div>
              </div>
              <small style="color:{MUTED};">%{pct:.1f}</small>
            </div>
            """, unsafe_allow_html=True)

    # Grafik
    with col_right:
        st.markdown('<div class="section-title">📈 Kategori Dağılımı</div>', unsafe_allow_html=True)
        cats = list(scope["kategori_toplamlari"].keys())
        vals = list(scope["kategori_toplamlari"].values())
        aktif = [(c, v) for c, v in zip(cats, vals) if v > 0]
        if aktif:
            try:
                import plotly.express as px
                fig = px.pie(
                    names=[c for c, _ in aktif],
                    values=[v for _, v in aktif],
                    hole=0.45,
                    color_discrete_sequence=["#1d6b45", "#2e8b57", "#3da873", "#d99a3d", "#c0573d", "#7f9c6b"],
                )
                fig.update_layout(
                    plot_bgcolor="white", paper_bgcolor="white",
                    margin=dict(t=10, b=10, l=0, r=0),
                    font=dict(family="Segoe UI", size=13),
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
                )
                fig.update_traces(textposition="inside", textinfo="percent")
                st.plotly_chart(fig, width='stretch')
            except Exception as e:
                st.error(f"Grafik oluşturulamadı: {e}")
        else:
            st.info("Grafik için yeterli veri yok.")

    st.markdown("---")

    # TEKİLLEŞTİRİLMİŞ KATEGORİ DETAYI
    st.markdown('<div class="section-title">🧾 Kategori & Alt Tür Detayları</div>', unsafe_allow_html=True)
    st.caption("Tüm tüketim ve emisyon verileri tek tabloda. TGA Tablo 10-13 çıktıları Excel/PDF dosyalarında yer alır.")

    detay_rows = []
    for kat, (emoji, birim) in {
        "Elektrik": ("🔌", "kWh"),
        "Doğal Gaz ve Yakıtlar": ("🔥", "m³/kg"),
        "Soğutucu & F-Gaz (Scope 1)": ("❄️", "kg"),
        "Araç Filosu & İş Seyahatleri (Scope 1/3)": ("🚗", "L/km"),
        "Su": ("🚿", "m³"),
        "Gıda Tüketimi": ("🍽️", "kg"),
        "Atık Yönetimi": ("♻️", "kg"),
        "Kimyasal Tüketimi": ("🧪", "L"),
    }.items():
        for alt_tur, miktar in tuketim.get(kat, {}).items():
            if not miktar or miktar <= 0:
                continue
            faktor = EMISSION_FACTORS.get(kat, {}).get(alt_tur, 0.0)
            emisyon = miktar * faktor
            pay = (emisyon / scope["toplam"] * 100) if scope["toplam"] > 0 else 0
            detay_rows.append({
                "Kategori": f"{emoji} {kat}",
                "Alt Tür": alt_tur,
                "Miktar": f"{miktar:,.1f} {birim}",
                "EF": faktor,
                "Emisyon (kg)": round(emisyon, 1),
                "Pay (%)": round(pay, 1),
            })

    if detay_rows:
        detay_df = pd.DataFrame(detay_rows).sort_values("Emisyon (kg)", ascending=False)
        st.dataframe(detay_df, width='stretch', hide_index=True)
    else:
        st.info("Henüz emisyon kaynağı bulunmuyor.")

    # Öne çıkan kaynaklar
    st.markdown('<div class="section-title">⚠️ Öne Çıkan Emisyon Kaynakları</div>', unsafe_allow_html=True)
    if en_agir:
        cols = st.columns(min(3, len(en_agir)))
        for sira, ((kat, deger, yuzde), col) in enumerate(zip(en_agir, cols), 1):
            with col:
                st.markdown(f"""
                <div class="data-card" style="text-align:center; border-top:4px solid {_color_for_pct(yuzde)};">
                  <div style="font-size:12px; color:{MUTED}; margin-bottom:4px;">#{sira} KAYNAK</div>
                  <div style="font-weight:800; font-size:17px;">{kat}</div>
                  <div style="color:{PRIMARY}; font-weight:800; font-size:24px;">{deger:,.0f}</div>
                  <div style="color:{MUTED}; font-size:12px;">kg CO₂e · toplamın %{yuzde}</div>
                </div>
                """, unsafe_allow_html=True)
    else:
        st.info("Henüz emisyon kaynağı bulunmuyor.")

    st.markdown("---")

    # Raporlar (raporlama alt sistemi)
    _rapor_karti(st.session_state.facility_id, period, r)

    st.markdown("---")

    # İndirmeler
    st.markdown('<div class="section-title">📥 İndirmeler</div>', unsafe_allow_html=True)
    col_xls, col_pdf, col_yeni = st.columns(3)
    with col_xls:
        try:
            import openpyxl  # noqa: F401
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            tablolar = tum_tablolar(r, period)
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                # Kapak sayfası – kurumsal kimlik
                kapak_df = pd.DataFrame([
                    ["KarbonAT P2 – TGA Uyumlu Takip Paketi"],
                    [f"Tesis: {tesis['ad']}"],
                    [f"Dönem: {format_donem(period)}"],
                    [f"Oluşturulma: {__import__('datetime').datetime.now().strftime('%d.%m.%Y %H:%M')}"],
                    [""],
                    ["Bu dosya KarbonAT tarafından tesisin gerçek aylık verilerinden üretilmiştir."],
                    ["GSTC / TGA Tablo 6,7,10-13 ve kimyasal envanter içerir."],
                    ["Her sayfa filtreli, yazdırma ayarlı ve denetime hazırdır."],
                ])
                kapak_df.to_excel(writer, sheet_name="Kapak", index=False, header=False)
                ws_kapak = writer.sheets["Kapak"]
                ws_kapak["A1"].font = Font(name="Calibri", size=14, bold=True, color="1D6B45")
                ws_kapak["A1"].alignment = Alignment(horizontal="left", vertical="center")
                for r in range(2, 9):
                    ws_kapak.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="17201c")
                ws_kapak.column_dimensions["A"].width = 78
                ws_kapak.sheet_properties.pageSetUpPr.fitToPage = True
                ws_kapak.page_setup.orientation = "portrait"
                ws_kapak.page_setup.paperSize = ws_kapak.PAPERSIZE_A4
                ws_kapak.page_setup.fitToWidth = 1
                ws_kapak.page_margins.left = 0.6
                ws_kapak.page_margins.right = 0.6
                for sheet, df in tablolar.items():
                    df.to_excel(writer, sheet_name=sheet[:31], index=False)
                    ws = writer.sheets[sheet[:31]]
                    # Kurumsal stil (raporlar.py ile aynı palet)
                    HEADER_FILL = PatternFill(start_color="1D6B45", end_color="1D6B45", fill_type="solid")
                    HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
                    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ZEBRA_FILL = PatternFill(start_color="EEF4EE", end_color="EEF4EE", fill_type="solid")
                    CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    THIN = Side(style="thin", color="CFD8CF")
                    thin_border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                    ncols = len(df.columns)
                    nrows = len(df) + 1
                    for c in range(1, ncols + 1):
                        cell = ws.cell(row=1, column=c)
                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT
                        cell.alignment = HEADER_ALIGN
                        cell.border = thin_border
                    for r in range(2, nrows + 1):
                        for c in range(1, ncols + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.alignment = CELL_ALIGN
                            cell.border = thin_border
                            cell.font = Font(name="Calibri", size=9 if ncols < 8 else 8)
                            if r % 2 == 1 and r >= 3:
                                cell.fill = ZEBRA_FILL
                            try:
                                if isinstance(cell.value, (int, float)):
                                    cell.alignment = Alignment(horizontal="right", vertical="center")
                            except Exception:
                                pass
                    for col_idx, col_name in enumerate(df.columns, 1):
                        max_len = len(str(col_name))
                        for val in df[col_name].astype(str):
                            if len(val) > max_len:
                                max_len = len(val)
                        width = min(36, max(12, max_len * 1.05 + 2))
                        if ncols >= 10:
                            width = min(width, 18)
                        elif ncols >= 7:
                            width = min(width, 24)
                        ws.column_dimensions[get_column_letter(col_idx)].width = width
                    ws.row_dimensions[1].height = 26 if ncols >= 8 else 20
                    for r in range(2, nrows + 1):
                        ws.row_dimensions[r].height = 16
                    ws.freeze_panes = "A2"
                    try:
                        ws.auto_filter.ref = ws.dimensions
                    except Exception:
                        pass
                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                    ws.page_setup.orientation = "landscape" if ncols >= 7 else "portrait"
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                    ws.page_margins.left = 0.4
                    ws.page_margins.right = 0.4
                    ws.page_margins.top = 0.5
                    ws.page_margins.bottom = 0.5
                    ws.print_title_rows = "1:1"
            st.download_button(
                label="📊 TGA Tabloları (Excel)",
                data=buffer.getvalue(),
                file_name=f"KarbonAT_{tesis['ad'].replace(' ', '_')}_{period}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch',
            )
        except ImportError:
            st.warning("Excel için `openpyxl` kurun: pip install openpyxl")

    with col_pdf:
        try:
            green_pdf = save_green_report(r, period, onceki)
            st.download_button(
                label="🌿 Yeşil Rapor (PDF)",
                data=green_pdf,
                file_name=f"KarbonAT_YeşilRapor_{tesis['ad'].replace(' ', '_')}_{period}.pdf",
                mime="application/pdf",
                width='stretch',
            )
        except Exception as e:
            st.error(f"PDF oluşturulamadı: {e}")

    with col_yeni:
        if st.button("🔄 Yeni Hesaplama", width='stretch'):
            st.session_state.step = 2
            st.session_state.tuketim = _sifir_tuketim()
            st.rerun()

    st.markdown("---")

    # Geçmiş
    st.markdown('<div class="section-title">📅 Geçmiş Kayıtlar</div>', unsafe_allow_html=True)
    if st.session_state.history:
        kayit_df = pd.DataFrame([
            {
                "Dönem": format_donem(k["period"]),
                "Toplam (ton)": k["sonuc"]["metrikler"]["toplam_ton"],
                "Oda-Gün (kg)": k["sonuc"]["metrikler"]["oda_gun_kg"],
                "Müşteri (kg)": k["sonuc"]["metrikler"]["musteri_kg"],
                "m² (kg)": k["sonuc"]["metrikler"]["m2_aylik_kg"],
            }
            for k in st.session_state.history
        ])
        st.dataframe(kayit_df, width='stretch', hide_index=True)
        try:
            import plotly.express as px
            fig2 = px.line(kayit_df, x="Dönem", y="Toplam (ton)", markers=True,
                           color_discrete_sequence=[PRIMARY])
            fig2.update_layout(plot_bgcolor="white", paper_bgcolor="white",
                               margin=dict(t=10, b=10, l=0, r=0),
                               font=dict(family="Segoe UI", size=13))
            st.plotly_chart(fig2, width='stretch')
        except Exception:
            st.bar_chart(kayit_df.set_index("Dönem")["Toplam (ton)"])
    else:
        st.caption("Her aylık hesaplama otomatik olarak kaydedilir ve burada trend oluşur.")

    st.markdown(
        '<div class="footer">KarbonAT P2 · GSTC / TGA Uyumlu · v0.4</div>',
        unsafe_allow_html=True,
    )


# ==============================================
# ADIM 5 - İÇERİK MERKEZİ (CONTENT ENGINE v2)
# ==============================================
def _aktif_sonuc(fac_id):
    """Oturumda hesaplama yoksa tesisin son kaydındaki gerçek veriyi kullanır."""
    sonuc = st.session_state.sonuc
    if isinstance(sonuc, dict) and sonuc:
        return sonuc
    kayitlar = list_records(fac_id)
    if kayitlar:
        s = kayitlar[-1].get("sonuc")
        if isinstance(s, dict):
            return s
    return None


def _icerik_detay(fac_id, tur):
    tur_id = tur["id"]
    defaults = varsayilan_tercih(tur_id)
    p = {**defaults, **(get_content_prefs(fac_id, tur_id) or {})}

    h1, h2 = st.columns([6, 1])
    with h1:
        st.markdown(f"### {tur['emoji']} {tur['baslik']}")
    with h2:
        grup = tur.get("grup", "")
        if st.button("✖ Kapat", key=f"mkapat_{fac_id}_{grup}"):
            st.session_state.pop(f"medya_detay_{fac_id}_{grup}", None)
            st.rerun()
    st.caption(tur["aciklama"])

    with st.expander("🗂️ Planlanan Yapı (iskelet)", expanded=False):
        for i, baslik in enumerate(ISKELETLER[tur_id], 1):
            st.markdown(f"**{i}.** {baslik}", unsafe_allow_html=True)

    st.markdown('<div class="section-title">🎨 Tasarım Tercihleri</div>', unsafe_allow_html=True)
    for soru in tercih_sorulari(tur_id):
        secenekler = soru["secenekler"]
        idx = secenekler.index(p.get(soru["anahtar"])) if p.get(soru["anahtar"]) in secenekler else 0
        p[soru["anahtar"]] = st.selectbox(
            soru["soru"],
            secenekler,
            index=idx,
            key=f"tp_{fac_id}_{tur_id}_{soru['anahtar']}",
        )

    tema_keys = list(TEMALAR.keys())
    tema_adlar = [TEMALAR[k]["ad"] for k in tema_keys]
    tema_sec = st.selectbox(
        "Görsel Tema",
        tema_adlar,
        index=tema_keys.index(p["tema"]) if p["tema"] in tema_keys else 0,
        key=f"tp_{fac_id}_{tur_id}_tema",
        help="Üretim aşamasında rapor/broşür görsel paletine uygulanacak.",
    )
    p["tema"] = tema_keys[tema_adlar.index(tema_sec)]

    p["notlar"] = st.text_area(
        "✍️ AI'a tarzı / kapsamı anlat (isteğe bağlı)",
        value=p.get("notlar", ""),
        placeholder=("Örn: 'Sıcak ve samimi bir dille, sayılarla değil hikayelerle anlat. "
                     "Yenilenebilir enerjiye geçişimizi öne çıkar.'"),
        key=f"tp_{fac_id}_{tur_id}_notlar",
    )

    save_content_prefs(fac_id, tur_id, p)

    st.markdown("---")
    aktif = st.button(
        f"🚀 {tur['baslik']} Üret",
        width='stretch',
        key=f"uretim_{fac_id}_{tur_id}",
        help="Tesis verileri + tercihlerin + TGA şablon referansları (RAG) ile AI içeriği üretir.",
    )
    if aktif:
        _t = st.session_state.tesis
        tesis = _t if isinstance(_t, dict) else {}
        _s = _aktif_sonuc(fac_id)
        sonuc = _s if isinstance(_s, dict) else None
        if not ai_engine.kbs_var_mi():
            st.info("Bilgi bankası (data/kb.json) yok. `python kb_build.py` ile oluştur. RAG'sız yine de üretilecek.")
        try:
            with st.spinner("AI içeriği üretiliyor..."):
                markdown = ai_engine.uretim_olustur(tur_id, tesis or {}, sonuc, p)
            save_media(fac_id, tur_id, markdown)
            st.session_state.pop(f"mte_{fac_id}_{tur_id}", None)
            st.toast("İçerik kaydedildi; düzenleyip PDF/Word indirebilirsin.")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Üretim başarısız: {e}")
            sy = str(e)
            if "429" in sy or "RESOURCE_EXHAUSTED" in sy or "quota" in sy.lower():
                st.warning(
                    "Kota (429) aşıldı — ücretsiz katmanın günlük/dakikalık limiti "
                    "tükendi. Dakika kotası birkaç saniyede açılır; üst üste deneme "
                    "engellemeyi uzatır. Kalıcı çözüm: GEMINI_API_KEY'i ücretli bir "
                    "anahtarla değiştirin. https://ai.google.dev/gemini-api/docs/rate-limits"
                )
                try:
                    fb = _ai_fallback_icerik(tur_id, tesis or {}, sonuc)
                    save_media(fac_id, tur_id, fb)
                    st.info("AI kotası dolu — geçici taslak kaydedildi. Birkaç dakika sonra yeniden deneyin.")
                    st.session_state.pop(f"mte_{fac_id}_{tur_id}", None)
                    st.rerun()
                except Exception:
                    pass
            else:
                st.caption("İpucu: GEMINI_API_KEY .env içinde olmalı; kota yoksa birkaç saniye sonra tekrar deneyin.")

    kayit = get_media(fac_id, tur_id)
    if not kayit:
        st.info("Henüz üretilmedi — yukarıdaki düğmeyle üretin; içerik düzenlenebilir ve PDF/Word indirilebilir.")
        return

    mte_key = f"mte_{fac_id}_{tur_id}"
    icerik = st.session_state.get(mte_key) or kayit["metin"]
    sonuc = _aktif_sonuc(fac_id)
    tasarim_var = tur_id in tasarim.TASARIMLAR

    # ---------- AI GÖRSEL ÜRETİMİ (gerçek fotoğraf/resim) ----------
    st.markdown("### 🖼️ AI Görseli")
    c1, c2 = st.columns([3, 2])
    with c1:
        gorsel_aktif = st.button(
            "🎨 AI Görsel Üret",
            key=f"gorsel_{fac_id}_{tur_id}",
            help="Tesis + tercihlerinle Gemini görsel modeli gerçek bir fotoğraf/resim üretir (şablon değil).",
            width='stretch',
        )
    with c2:
        st.caption("Şablondan bağımsız; tema, vurgu ve notlar kullanılır.")
    if gorsel_aktif:
        try:
            with st.spinner("AI görseli üretiliyor (görsel modelleri ayrı kota kullanır)..."):
                gorsel_bytes = ai_engine.gorsel_uret(tur_id, st.session_state.tesis or {}, sonuc or None, p)
            save_media(fac_id, tur_id, icerik, gorsel=gorsel_bytes)
            st.toast("AI görseli üretildi ve kaydedildi.")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Görsel üretimi başarısız: {e}")

    gorsel_yol = kayit.get("gorsel_yol")
    if gorsel_yol and os.path.exists(gorsel_yol):
        with open(gorsel_yol, "rb") as f:
            gorsel_bytes = f.read()
        st.image(gorsel_bytes, width='stretch')
        st.download_button(
            "⬇️ AI Görseli İndir (PNG)",
            data=gorsel_bytes,
            file_name=f"KarbonAT_{tur_id}_{fac_id}_ai.png",
            mime="image/png",
            key=f"mga_{fac_id}_{tur_id}",
            width='stretch',
        )
    else:
        st.caption("Henüz AI görseli yok — yukarıdaki düğmeyle üretin.")

    st.markdown("---")

    if tasarim_var:
        t_duzen = tasarim.TASARIMLAR[tur_id]
        st.markdown("### 🎨 Görsel Çıktı")
        if not sonuc:
            st.caption("Not: Bu tesis için henüz hesaplama kaydı yok; sayı kartları boş görünür. Önce adım 4'te hesaplama yapın.")
            sonuc = {}
        html = getattr(tasarim, t_duzen["html"])(sonuc, icerik)
        if "png" in t_duzen:
            png = getattr(tasarim, t_duzen["png"])(sonuc, icerik, tur_id)
            st.image(png, width='stretch')
            with st.expander("Ayrıca web / HTML görünümü", expanded=False):
                st.html(html, height=520)
        else:
            with st.expander("Canlı önizleme", expanded=True):
                st.html(html, height=660)

        with st.expander("✏️ Ham içeriği düzenle (markdown)", expanded=False):
            st.text_area("İçerik", value=icerik, height=380, key=mte_key)
            st.caption("Başlık: `### Başlık`  ·  Madde: `- metin`. Düzenleme görsele/PDF/HTML'e yansır.")
        icerik = st.session_state.get(mte_key) or kayit["metin"]

        st.markdown("---")
        n_buton = 5 if "png" in t_duzen else 4
        cols = st.columns(n_buton)
        with cols[0]:
            try:
                st.download_button("🖼️ PDF", data=getattr(tasarim, tasarim.TASARIMLAR[tur_id]["pdf"])(sonuc, icerik),
                                   file_name=f"KarbonAT_{tur_id}_{fac_id}.pdf", mime="application/pdf",
                                   key=f"mpdf_{fac_id}_{tur_id}", width='stretch')
            except Exception as e:  # noqa: BLE001
                st.warning(f"PDF: {e}")
        if "png" in t_duzen:
            with cols[1]:
                st.download_button("🏞️ PNG", data=png,
                                   file_name=f"KarbonAT_{tur_id}_{fac_id}.png",
                                   mime="image/png",
                                   key=f"mpng_{fac_id}_{tur_id}", width='stretch')
            with cols[2]:
                st.download_button("🌐 HTML", data=html.encode("utf-8"),
                                   file_name=f"KarbonAT_{tur_id}_{fac_id}.html", mime="text/html",
                                   key=f"mthtml_{fac_id}_{tur_id}", width='stretch')
            with cols[3]:
                try:
                    st.download_button("📄 Word", data=raporlar.rapor_docx(icerik),
                                       file_name=f"KarbonAT_{tur_id}_{fac_id}.docx",
                                       mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                       key=f"mdoc_{fac_id}_{tur_id}", width='stretch')
                except Exception as e:  # noqa: BLE001
                    st.warning(f"Word: {e}")
            with cols[4]:
                if st.button("💾 Kaydet", key=f"msav_{fac_id}_{tur_id}", width='stretch'):
                    save_media(fac_id, tur_id, icerik)
                    st.toast("Düzenleme kaydedildi.")
                    st.rerun()
        else:
            with cols[1]:
                st.download_button("🌐 HTML", data=html.encode("utf-8"),
                                   file_name=f"KarbonAT_{tur_id}_{fac_id}.html", mime="text/html",
                                   key=f"mthtml_{fac_id}_{tur_id}", width='stretch')
            with cols[2]:
                try:
                    st.download_button("📄 Word", data=raporlar.rapor_docx(icerik),
                                       file_name=f"KarbonAT_{tur_id}_{fac_id}.docx",
                                       mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                       key=f"mdoc_{fac_id}_{tur_id}", width='stretch')
                except Exception as e:  # noqa: BLE001
                    st.warning(f"Word: {e}")
            with cols[3]:
                if st.button("💾 Kaydet", key=f"msav_{fac_id}_{tur_id}", width='stretch'):
                    save_media(fac_id, tur_id, icerik)
                    st.toast("Düzenleme kaydedildi.")
                    st.rerun()
        return

    icerik = st.text_area(
        "İçerik — buradan düzenleyebilirsin",
        value=kayit["metin"],
        height=300,
        key=mte_key,
    )
    st.caption("Tablolar aşağıda kaydırmalı tablo olarak gösterilir; PDF/Word gerçek tablo içerir.")
    _markdown_blok(icerik)

    st.markdown("---")
    d1, d2, d3 = st.columns(3)
    with d1:
        try:
            st.download_button("⬇️ PDF", data=raporlar.rapor_pdf(icerik),
                               file_name=f"KarbonAT_{tur_id}_{fac_id}.pdf", mime="application/pdf",
                               key=f"mpdf_{fac_id}_{tur_id}", width='stretch')
        except Exception as e:  # noqa: BLE001
            st.warning(f"PDF: {e}")
    with d2:
        try:
            st.download_button("📄 Word (.docx)", data=raporlar.rapor_docx(icerik),
                               file_name=f"KarbonAT_{tur_id}_{fac_id}.docx",
                               mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                               key=f"mdoc_{fac_id}_{tur_id}", width='stretch')
        except Exception as e:  # noqa: BLE001
            st.warning(f"Word: {e}")
    with d3:
        if st.button("💾 Düzenlemeyi Kaydet", key=f"msav_{fac_id}_{tur_id}", width='stretch'):
            save_media(fac_id, tur_id, icerik)
            st.toast("Düzenleme kaydedildi.")
            st.rerun()


def _icerik_grid(fac_id, grup_id, turler):
    secim_key = f"medya_detay_{fac_id}_{grup_id}"
    SUTUN = 3
    cols = st.columns(SUTUN)
    for i, tur in enumerate(turler):
        kayit = get_media(fac_id, tur["id"])
        durum = "✅ Kayıtlı" if kayit else "⏳ Beklemede"
        with cols[i % SUTUN]:
            if st.button(f"{tur['emoji']} {tur['baslik']}", key=f"mcd_{fac_id}_{grup_id}_{tur['id']}",
                         width='stretch'):
                st.session_state[secim_key] = tur["id"]
                st.rerun()
            st.caption(f"{durum} · {' · '.join(tur.get('ciktilar', []))}")

    secilen = st.session_state.get(secim_key)
    if secilen and secilen in {t["id"] for t in turler}:
        st.markdown("---")
        tur = next(t for t in turler if t["id"] == secilen)
        _icerik_detay(fac_id, tur)


def adim_icerik():
    st.markdown(mvp_stepper(active=4), unsafe_allow_html=True)
    tesis = st.session_state.tesis
    if not tesis or not tesis.get("id"):
        st.warning("Önce bir tesis seçin.")
        return

    st.markdown('<h1>📣 Medya & İçerik</h1>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="chip">🌿 {tesis["ad"]}</div>'
        f'<div class="chip">📣 Medya üretimi</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="banner"><strong>Medya &amp; İçerik</strong>: web sayfası, broşür, QR/oda kartı, '
        'basın bülteni, sosyal medya, görsel/afiş, eğitim ve anket şablonları. '
        '<strong>Raporlar</strong> (Tablo 1-13, politikalar) ayrıdır ve sonuç ekranında / profilinizde '
        'yer alır. Her tür yalnızca kendi başlığına yönelik içerik üretir.</div>',
        unsafe_allow_html=True,
    )

    # Amaca göre gruplar (AMAC_GRUPLARI); yalnız medya türleri gösterilir
    gruplar = []
    for g in AMAC_GRUPLARI:
        turler = [t for t in ICERIK_TURLERI
                  if t.get("grup") == g["id"] and t.get("sistem") == "medya"]
        if not turler:
            continue
        gruplar.append({**g, "turler": turler})

    tabs = st.tabs([f'{g["emoji"]} {g["baslik"]}' for g in gruplar])
    for tab, g in zip(tabs, gruplar):
        with tab:
            # Alt grup (örn. anket) paylaşanlar alt sekmelerde, tek olanlar kart olarak
            yiginlar = {}
            for t in g["turler"]:
                yiginlar.setdefault(t.get("alt_grup", t["id"]), []).append(t)
            for yig in yiginlar.values():
                if len(yig) == 1:
                    _icerik_grid(tesis["id"], g["id"], yig)
                else:
                    alt_tabs = st.tabs([t["alt_baslik"] for t in yig])
                    for at, t in zip(alt_tabs, yig):
                        with at:
                            _icerik_grid(tesis["id"], g["id"], [t])

    st.markdown("---")
    if st.button("← Tesis Seçimine Dön", type="secondary"):
        st.session_state.step = 0
        st.rerun()


# ==============================================
# ROUTER
# ==============================================
if st.session_state.user is None:
    auth_screen()
else:
    sidebar()

    adimler = [
        adim_tesis_secimi,
        adim_tesis,
        adim_veri,
        adim_hesap,
        adim_sonuc,
        adim_icerik,
    ]

    adimler[st.session_state.step]()